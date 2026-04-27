#!/usr/bin/env python3
"""Build a safe local Firestore import preview from the PNB Excel workbooks.

The script never writes to Firebase. It produces a JSON file that an admin can
review in the web app's "PNB İçe Aktar" tab before committing to Firestore.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ID = "pnb"
PROJECT_TITLE = "Pertev Naili Boratav Arşivi Dijitalleştirme"


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    replacements = {
        "ı": "i",
        "İ": "I",
        "ş": "s",
        "Ş": "S",
        "ğ": "g",
        "Ğ": "G",
        "ü": "u",
        "Ü": "U",
        "ö": "o",
        "Ö": "O",
        "ç": "c",
        "Ç": "C",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def slug(value: Any, fallback: str = "item") -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or fallback


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def to_number(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def to_int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


# Series-no helpers for priority and suitableFor classification.
# A series cell may contain a slash-list ("120.1/110.2/..."); we use the first
# segment for root-aware checks but keep the full string for prefix checks.

def _series_first_segment(series_no: str) -> str:
    return series_no.split("/")[0].strip() if series_no else ""


def derive_priority(series_no: str) -> str:
    series = clean_string(series_no)
    first = _series_first_segment(series)
    normalized = first.rstrip(".")
    if first.startswith("120"):
        return "high"
    if normalized == "170.3":
        return "high"
    if first.startswith("19.1"):
        return "medium"
    medium_roots = ("110", "12", "13", "16", "18")
    for root in medium_roots:
        if normalized == root or first.startswith(root + "."):
            return "medium"
    return "low"


def derive_suitable_for(series_no: str, material_type: str) -> list[str]:
    tags: list[str] = []
    series = clean_string(series_no)
    material = clean_string(material_type)
    first = _series_first_segment(series)
    if series.startswith("170"):
        tags.append("osmanlica")
    if (
        material.startswith("Dia")
        or material.startswith("Negatif")
        or material.startswith("I.01")
        or series.startswith("220")
    ):
        tags.append("visual")
    if material == "K" and first.startswith("120.5"):
        tags.append("transcription")
    if not tags:
        tags.append("general")
    return tags


def _split_count(total: int | None, n: int, index: int) -> int | None:
    if total is None:
        return None
    base = total // n
    if index == n - 1:
        return total - base * (n - 1)
    return base


def split_unit_if_needed(unit: dict[str, Any], threshold: int) -> list[dict[str, Any]]:
    pages = unit.get("pageCount")
    if pages is None or pages <= threshold:
        return [unit]
    n = (pages + threshold - 1) // threshold
    description = unit.get("contentDescription") or ""
    base_pages = unit.get("pageCount")
    base_docs = unit.get("documentCount")
    base_folders = unit.get("folderCount")
    base_files = unit.get("fileCount") or 0  # legacy field, 0-default
    box_no = unit.get("boxNo", "")
    source_code = unit.get("sourceCode", "")
    series_no = unit.get("seriesNo", "")
    sub_units: list[dict[str, Any]] = []
    for index in range(n):
        letter = chr(ord("A") + index)
        sub_box = f"{box_no}{letter}"
        sub_pages = _split_count(base_pages, n, index)
        sub_docs = _split_count(base_docs, n, index)
        sub_folders = _split_count(base_folders, n, index)
        # legacy fileCount: keep 0-default semantics
        sub_files = _split_count(base_files, n, index) or 0
        range_indicator = f"Kısım {letter} ({index + 1}/{n})"
        sub_description = f"{description} · {range_indicator}".strip(" ·") if description else range_indicator
        sub = dict(unit)
        sub["id"] = f"{unit['id']}-{letter.lower()}"
        sub["boxNo"] = sub_box
        sub["title"] = f"PNB {source_code} / {series_no} - Kutu {sub_box}"
        sub["pageCount"] = sub_pages
        sub["documentCount"] = sub_docs
        sub["folderCount"] = sub_folders
        sub["fileCount"] = sub_files
        sub["contentDescription"] = sub_description
        sub["notes"] = sub_description
        sub["sourceIdentifier"] = f"{source_code} / {series_no} / {sub_box}".strip()
        sub["splitParentId"] = unit["id"]
        sub["splitIndex"] = index + 1
        sub["splitTotal"] = n
        sub_units.append(sub)
    return sub_units


def to_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return clean_string(value) or None


def split_people(value: Any) -> list[str]:
    text = clean_string(value)
    if not text:
        return []
    parts = re.split(r"[,;\n/&]+", text)
    return [clean_string(part) for part in parts if clean_string(part)]


def find_file(excel_dir: Path, *needles: str) -> Path:
    normalized_needles = [normalize_text(needle) for needle in needles]
    for path in excel_dir.glob("*.xlsx"):
        name = normalize_text(path.name)
        if all(needle in name for needle in normalized_needles):
            return path
    available = ", ".join(path.name for path in excel_dir.glob("*.xlsx"))
    raise FileNotFoundError(f"Could not find workbook matching {needles}. Available: {available}")


def rows_as_dicts(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_string(value) for value in rows[0]]
    normalized_headers = [normalize_text(header) for header in headers]
    result: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(clean_string(value) for value in row):
            continue
        record = {"_sourceRow": row_index}
        for header, normalized, value in zip(headers, normalized_headers, row):
            if not normalized:
                continue
            record[normalized] = value
            record[f"_label_{normalized}"] = header
        result.append(record)
    return result


def value(record: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        normalized = normalize_text(key)
        if normalized in record:
            return record[normalized]
    return None


def build_people(path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, str]]]:
    people: list[dict[str, Any]] = []
    seen: Counter[str] = Counter()
    rows = rows_as_dicts(path, "Sayfa1")
    for row in rows:
        full_name = clean_string(value(row, "paydaş", "paydas"))
        if not full_name:
            continue
        email = clean_string(value(row, "email")).lower()
        person = {
            "id": f"pnb-person-{slug(email or full_name)}",
            "projectId": PROJECT_ID,
            "fullName": full_name,
            "normalizedName": normalize_text(full_name),
            "email": email,
            "phone": clean_string(value(row, "telefon")),
            "profession": clean_string(value(row, "meslek")),
            "university": clean_string(value(row, "universite", "üniversite")),
            "educationDepartment": clean_string(value(row, "bölüm", "bolum")),
            "city": clean_string(value(row, "şehir", "sehir")),
            "foundationRole": clean_string(value(row, "tarih vakfındaki rolü", "tarih vakfindaki rolu")),
            "projectRole": clean_string(value(row, "projede rolü", "projede rolu")),
            "expectation": clean_string(value(row, "projeden beklentisi")),
            "power": to_number(value(row, "gücü (1-5)", "gucu (1-5)")),
            "interest": to_number(value(row, "ilgisi (1-5)")),
            "stakeholderLevel": clean_string(value(row, "güç/ilgi seviyesi", "guc/ilgi seviyesi")),
            "sourceRow": row["_sourceRow"],
        }
        seen[person["normalizedName"]] += 1
        people.append(person)
    by_name = {person["normalizedName"]: person for person in people}
    duplicates = [
        {"name": name, "count": count}
        for name, count in sorted(seen.items())
        if name and count > 1
    ]
    return people, by_name, duplicates


def build_availability(path: Path, people_by_name: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sayfa1"]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean_string(value) for value in rows[0]]
    day_map = {
        "pztesi": "monday",
        "pazartesi": "monday",
        "sali": "tuesday",
        "salı": "tuesday",
        "cars": "wednesday",
        "carsamba": "wednesday",
        "çarş": "wednesday",
        "per": "thursday",
        "persembe": "thursday",
        "cuma": "friday",
    }
    availability: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(clean_string(cell) for cell in row):
            continue
        full_name = clean_string(row[0])
        if not full_name:
            continue
        normalized_name = normalize_text(full_name)
        matched_person = people_by_name.get(normalized_name)
        slots = []
        for col_index, cell in enumerate(row[3:13], start=3):
            if normalize_text(cell) != "x":
                continue
            label = headers[col_index] if col_index < len(headers) else f"slot-{col_index}"
            normalized_label = normalize_text(label)
            day = next((mapped for key, mapped in day_map.items() if key in normalized_label), "unknown")
            slot_no = "2" if "2" in normalized_label else "1"
            slots.append({"label": label, "day": day, "slot": slot_no})
        availability.append(
            {
                "id": f"pnb-availability-{slug(full_name)}",
                "projectId": PROJECT_ID,
                "personName": full_name,
                "normalizedName": normalized_name,
                "email": matched_person.get("email", "") if matched_person else "",
                "userUid": "",
                "topics": [clean_string(row[1]), clean_string(row[2])],
                "slots": slots,
                "slotCount": len(slots),
                "sourceRow": row_index,
            }
        )
    return availability


def build_archive_units(
    path: Path,
    people_by_name: dict[str, dict[str, Any]],
    split_over: int = 500,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = rows_as_dicts(path)
    units: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    used_ids: Counter[str] = Counter()
    for row in rows:
        source_code = clean_string(value(row, "kaynak kodu"))
        series_no = clean_string(value(row, "seri no"))
        box_no = clean_string(value(row, "kutu no"))
        # Footnote rows (no series AND no box) are explanatory text on the
        # spreadsheet, not actual archive units. Drop them but record why.
        if not series_no and not box_no:
            skipped.append({
                "sourceRow": row["_sourceRow"],
                "reason": "no series and no box (footnote/comment row)",
                "note": clean_string(value(row, "not")),
            })
            continue
        base_id = f"pnb-{slug(source_code, 'source')}-{slug(series_no, 'series')}-{slug(box_no, 'box')}"
        used_ids[base_id] += 1
        unit_id = base_id if used_ids[base_id] == 1 else f"{base_id}-{used_ids[base_id]}"
        assigned_names = split_people(value(row, "atanan gönüllü(ler)", "atanan gonullu(ler)"))
        assigned_emails = []
        for name in assigned_names:
            person = people_by_name.get(normalize_text(name))
            if person and person.get("email"):
                assigned_emails.append(person["email"])
        material_type = clean_string(value(row, "materyal türü", "materyal turu"))
        content_description = clean_string(value(row, "not"))
        page_count = to_int_or_none(value(row, "sayfa adedi"))
        document_count = to_int_or_none(value(row, "belge adedi"))
        folder_count = to_int_or_none(value(row, "dosya adedi"))
        source_identifier = " / ".join(
            part for part in [source_code, series_no, box_no] if part
        )
        unit = {
            "id": unit_id,
            "projectId": PROJECT_ID,
            "projectTitle": PROJECT_TITLE,
            "title": f"PNB {source_code} / {series_no} - Kutu {box_no}",
            "sourceCode": source_code,
            "seriesNo": series_no,
            "boxNo": box_no,
            # Legacy 0-default counters (kept so existing dashboards keep summing).
            "fileCount": to_number(value(row, "dosya adedi")),
            "documentCount": document_count if document_count is not None else 0,
            "pageCount": page_count if page_count is not None else 0,
            # New null-aware counters per the report-first schema.
            "folderCount": folder_count,
            "materialType": material_type,
            "notes": content_description,
            "contentDescription": content_description,
            "sourceIdentifier": source_identifier,
            "priority": derive_priority(series_no),
            "suitableFor": derive_suitable_for(series_no, material_type),
            "city": "istanbul",
            "digitized": False,
            "startDate": to_date(value(row, "başlangıç tarihi", "baslangic tarihi")),
            "updatedDate": to_date(value(row, "güncelleme tarihi", "guncelleme tarihi")),
            "endDate": to_date(value(row, "bitiş tarihi", "bitis tarihi")),
            "assignedNames": assigned_names,
            "assignedToEmails": sorted(set(assigned_emails)),
            "assignedToUids": [],
            "completedFileCount": to_number(value(row, "tamamlanan dosya")),
            "completedDocumentCount": to_number(value(row, "tamamlanan belge adedi")),
            "completedPageCount": to_number(value(row, "tamamlanan sayfa adedi")),
            "remainingFileCount": to_number(value(row, "kalan dosya")),
            "remainingDocumentCount": to_number(value(row, "kalan belge adedi")),
            "remainingPageCount": to_number(value(row, "kalan sayfa adedi")),
            "status": "not_started",
            "blockerNote": "",
            "dueDate": None,
            "latestReportAt": None,
            "sourceRow": row["_sourceRow"],
        }
        # Preserve the new null-aware pageCount/documentCount on the unit so
        # split logic uses the correct totals; legacy 0-default fields above
        # are for downstream summing only.
        unit["pageCount"] = page_count if page_count is not None else 0
        unit["documentCount"] = document_count if document_count is not None else 0
        if page_count is None:
            unit["pageCount"] = None
        if document_count is None:
            unit["documentCount"] = None
        units.extend(split_unit_if_needed(unit, split_over))
    return units, skipped


def build_communication_plans(path: Path) -> list[dict[str, Any]]:
    rows = rows_as_dicts(path)
    plans: list[dict[str, Any]] = []
    for row in rows:
        title = clean_string(value(row, "iletişim tipi", "iletisim tipi"))
        if not title:
            continue
        plans.append(
            {
                "id": f"pnb-communication-{slug(title)}",
                "projectId": PROJECT_ID,
                "title": title,
                "goal": clean_string(value(row, "iletişimin hedefi", "iletisimin hedefi")),
                "channel": clean_string(value(row, "ortam")),
                "frequency": clean_string(value(row, "sıklık", "siklik")),
                "meetingPlan": clean_string(value(row, "toplantı planı", "toplabtı planı", "toplabti plani")),
                "participants": clean_string(value(row, "katılımcılar", "katilimcilar")),
                "owner": clean_string(value(row, "sahibi")),
                "deliverables": clean_string(value(row, "teslimatlar")),
                "format": clean_string(value(row, "format")),
                "sourceRow": row["_sourceRow"],
            }
        )
    return plans


def _archive_summary(units: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "archiveUnits": len(units),
        "fileCount": sum(unit.get("fileCount") or 0 for unit in units),
        "documentCount": sum(unit.get("documentCount") or 0 for unit in units),
        "pageCount": sum(unit.get("pageCount") or 0 for unit in units),
        "completedFileCount": sum(unit.get("completedFileCount") or 0 for unit in units),
        "completedDocumentCount": sum(unit.get("completedDocumentCount") or 0 for unit in units),
        "completedPageCount": sum(unit.get("completedPageCount") or 0 for unit in units),
    }


def build_preview(excel_dir: Path, split_over: int = 500) -> dict[str, Any]:
    stakeholder_file = find_file(excel_dir, "paydas")
    availability_file = find_file(excel_dir, "gonullu", "zaman")
    work_plan_file = find_file(excel_dir, "is", "plani")
    communication_file = find_file(excel_dir, "iletisim", "matrisi")

    people, people_by_name, duplicate_names = build_people(stakeholder_file)
    availability = build_availability(availability_file, people_by_name)
    archive_units, skipped_rows = build_archive_units(work_plan_file, people_by_name, split_over)
    communication_plans = build_communication_plans(communication_file)

    missing_emails = [
        {"name": person["fullName"], "sourceRow": person["sourceRow"]}
        for person in people
        if not person["email"]
    ]
    unmatched_availability = [
        {"name": item["personName"], "slotCount": item["slotCount"], "sourceRow": item["sourceRow"]}
        for item in availability
        if not item["email"]
    ]
    empty_assigned = [unit["id"] for unit in archive_units if not unit["assignedNames"]]

    summary = {
        **_archive_summary(archive_units),
        "people": len(people),
        "peopleWithEmail": sum(1 for person in people if person["email"]),
        "availabilityRows": len(availability),
        "availablePeople": sum(1 for item in availability if item["slotCount"] > 0),
        "availabilitySlotCount": sum(item["slotCount"] for item in availability),
        "communicationPlans": len(communication_plans),
    }

    return {
        "version": 1,
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "project": {"id": PROJECT_ID, "title": PROJECT_TITLE},
        "sourceFiles": {
            "stakeholders": stakeholder_file.name,
            "availability": availability_file.name,
            "workPlan": work_plan_file.name,
            "communication": communication_file.name,
        },
        "splitOver": split_over,
        "summary": summary,
        "archiveUnits": archive_units,
        "people": people,
        "availability": availability,
        "communicationPlans": communication_plans,
        "checks": {
            "missingEmails": missing_emails,
            "duplicateNames": duplicate_names,
            "unmatchedAvailabilityNames": unmatched_availability,
            "emptyAssignedArchiveUnits": empty_assigned,
            "skippedRows": skipped_rows,
        },
    }


def build_preview_work_plan_only(workbook: Path, split_over: int = 500) -> dict[str, Any]:
    """Single-file mode: only refresh archiveUnits.

    People, availability and communication plans are emitted as empty arrays so
    the existing Bakım flow can still consume the JSON, but the maintainer must
    be aware that committing this preview will reset the projects/pnb counters
    for those collections to 0 (peopleCount, availabilitySlotCount,
    communicationPlanCount). Use the full `build_preview` when those counters
    need to stay accurate.
    """
    archive_units, skipped_rows = build_archive_units(workbook, {}, split_over)
    summary = {
        **_archive_summary(archive_units),
        "people": 0,
        "peopleWithEmail": 0,
        "availabilityRows": 0,
        "availablePeople": 0,
        "availabilitySlotCount": 0,
        "communicationPlans": 0,
    }
    return {
        "version": 1,
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "project": {"id": PROJECT_ID, "title": PROJECT_TITLE},
        "mode": "work_plan_only",
        "sourceFiles": {"workPlan": workbook.name},
        "splitOver": split_over,
        "summary": summary,
        "archiveUnits": archive_units,
        "people": [],
        "availability": [],
        "communicationPlans": [],
        "checks": {
            "missingEmails": [],
            "duplicateNames": [],
            "unmatchedAvailabilityNames": [],
            "emptyAssignedArchiveUnits": [unit["id"] for unit in archive_units if not unit["assignedNames"]],
            "skippedRows": skipped_rows,
        },
    }


def write_summary_markdown(preview: dict[str, Any], output: Path) -> None:
    units = preview.get("archiveUnits", [])
    skipped = preview.get("checks", {}).get("skippedRows", [])
    split_over = preview.get("splitOver", 500)

    by_priority: Counter[str] = Counter()
    by_tag: Counter[str] = Counter()
    by_material: Counter[str] = Counter()
    parent_ids: set[str] = set()
    for unit in units:
        by_priority[unit.get("priority", "low")] += 1
        for tag in unit.get("suitableFor") or []:
            by_tag[tag] += 1
        material = unit.get("materialType") or "(boş)"
        by_material[material] += 1
        parent = unit.get("splitParentId")
        if parent:
            parent_ids.add(parent)
    split_sub_units = sum(1 for unit in units if unit.get("splitParentId"))

    lines: list[str] = []
    lines.append(f"# PNB import özeti")
    lines.append("")
    lines.append(f"- Üretim zamanı: {preview.get('generatedAt', '')}")
    lines.append(f"- Kaynak dosya: `{preview.get('sourceFiles', {}).get('workPlan', '-')}`")
    lines.append(f"- Bölme eşiği (Sayfa Adedi): {split_over}")
    lines.append(f"- Toplam birim (bölünmüş alt-birimler dahil): **{len(units)}**")
    lines.append(f"- Bölünen üst birim: {len(parent_ids)} → {split_sub_units} alt-birim")
    lines.append("")
    lines.append("## Öncelik dağılımı")
    for level in ("high", "medium", "low"):
        lines.append(f"- {level}: {by_priority.get(level, 0)}")
    lines.append("")
    lines.append("## Uygunluk etiketleri (suitableFor)")
    for tag, count in sorted(by_tag.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"- `{tag}`: {count}")
    lines.append("")
    lines.append("## Materyal türü")
    for material, count in sorted(by_material.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"- `{material}`: {count}")
    lines.append("")
    lines.append("## Atlanan satırlar")
    if not skipped:
        lines.append("Yok.")
    else:
        for entry in skipped:
            note = entry.get("note") or ""
            note_short = (note[:120] + "…") if len(note) > 120 else note
            lines.append(f"- satır {entry.get('sourceRow')}: {entry.get('reason')} — {note_short}")
    lines.append("")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a PNB Firebase import preview JSON.")
    parser.add_argument("--excel-dir", type=Path, default=Path(".."), help="Directory containing the four PNB Excel files (full mode).")
    parser.add_argument("--workbook", type=Path, default=None, help="Path to a single PNB iş planı workbook for work-plan-only mode.")
    parser.add_argument("--split-over", type=int, default=500, help="Split any archive unit whose Sayfa Adedi exceeds this threshold into A/B/C sub-units.")
    parser.add_argument("--output", type=Path, default=Path("imports/pnb-import-preview.json"), help="Output JSON path.")
    parser.add_argument("--summary", type=Path, default=None, help="Optional human-readable Markdown summary path.")
    args = parser.parse_args()

    if args.workbook:
        preview = build_preview_work_plan_only(args.workbook.resolve(), args.split_over)
    else:
        preview = build_preview(args.excel_dir.resolve(), args.split_over)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.output}")
    if args.summary:
        write_summary_markdown(preview, args.summary)
        print(f"Wrote {args.summary}")
    print(json.dumps(preview["summary"], ensure_ascii=False, indent=2))
    checks = preview["checks"]
    print(
        "Checks: "
        f"{len(checks['missingEmails'])} missing emails, "
        f"{len(checks['duplicateNames'])} duplicate names, "
        f"{len(checks['unmatchedAvailabilityNames'])} unmatched availability rows, "
        f"{len(checks['emptyAssignedArchiveUnits'])} unassigned archive units, "
        f"{len(checks.get('skippedRows', []))} skipped rows."
    )


if __name__ == "__main__":
    main()
