"""Tarih Vakfı Kütüphanesi — kitap envanteri şablonu üretir."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

BORDO, ACIK, KREM = "601040", "F3EFF1", "FBF2F7"
SATIR = 3000  # doğrulama ve formüllerin kapsadığı satır sayısı

wb = Workbook()

# ═══════════════ 1) BAŞLA ═══════════════
b = wb.active
b.title = "Başla"

b["A1"] = "Tarih Vakfı Kütüphanesi — Kitap Envanteri"
b["A1"].font = Font(name="Arial", bold=True, size=16, color=BORDO)
b["A2"] = "Taşınma ve ayıklama çalışması"
b["A2"].font = Font(name="Arial", size=10, color="74686E")

adimlar = [
    ("NASIL KULLANILIR", None),
    ("1. Aşağıdaki sarı hücrelere ölçüm sonuçlarını yazın.", None),
    ("2. Kayıtları 'Envanter' sayfasına girin. Gri başlıklı sütunlar elle doldurulur.", None),
    ("3. Kategori, Kural kodu ve Durum sütunlarında hücreye tıklayınca liste açılır.", None),
    ("4. Kutulama sırasında 'Kutular' sayfasını doldurun, kitap satırlarına kutu no yazın.", None),
    ("5. 'Özet' sayfası kendiliğinden hesaplanır; elle bir şey yazmayın.", None),
]
r = 4
for metin, _ in adimlar:
    b.cell(r, 1, metin)
    if metin.isupper():
        b.cell(r, 1).font = Font(name="Arial", bold=True, size=11, color=BORDO)
    else:
        b.cell(r, 1).font = Font(name="Arial", size=10.5)
    r += 1

# Ölçüm girdileri (sarı = doldurulacak)
r += 1
b.cell(r, 1, "ÖLÇÜMLER — bu sarı hücreleri siz doldurun").font = Font(name="Arial", bold=True, size=11, color=BORDO)
r += 1
olcumler = [
    ("Eski binadaki dolu raf uzunluğu (metre)", 200, "Dolu rafları metreyle ölçüp toplayın"),
    ("Yeni binadaki raf uzunluğu (metre)", 110, "Yeni binada kitaba ayrılan raf"),
    ("Metre başına ortalama kitap", 32, "Bir metre rafta kaç kitap duruyor (sayarak bulun; tipik 30–35)"),
]
ilk_olcum = r
for ad, deger, aciklama in olcumler:
    b.cell(r, 1, ad).font = Font(name="Arial", size=10.5)
    h = b.cell(r, 2, deger)
    h.font = Font(name="Arial", bold=True, size=11, color="0000FF")
    h.fill = PatternFill("solid", fgColor="FFFF99")
    h.alignment = Alignment(horizontal="center")
    h.border = Border(*[Side(style="thin", color="BBBBBB")] * 4)
    b.cell(r, 3, aciklama).font = Font(name="Arial", size=9, color="74686E")
    r += 1

r += 1
b.cell(r, 1, "HESAPLANAN").font = Font(name="Arial", bold=True, size=11, color=BORDO)
r += 1
hesaplar = [
    ("Tahmini toplam kitap", f"=B{ilk_olcum}*B{ilk_olcum+2}"),
    ("Yeni binaya sığacak kitap", f"=B{ilk_olcum+1}*B{ilk_olcum+2}"),
    ("Ayıklanması gereken kitap", f"=B{ilk_olcum}*B{ilk_olcum+2}-B{ilk_olcum+1}*B{ilk_olcum+2}"),
    ("Ayıklama oranı", f"=IF(B{ilk_olcum}=0,0,1-B{ilk_olcum+1}/B{ilk_olcum})"),
]
ilk_hesap = r
for ad, form in hesaplar:
    b.cell(r, 1, ad).font = Font(name="Arial", size=10.5)
    h = b.cell(r, 2, form)
    h.font = Font(name="Arial", bold=True, size=11)
    h.alignment = Alignment(horizontal="center")
    h.number_format = "0%" if "oran" in ad.lower() else "#,##0"
    r += 1

# Örnek satır — Envanter'e nasıl girileceğini gösterir
r += 2
b.cell(r, 1, "ÖRNEK KAYIT — 'Envanter' sayfasında bir satır böyle görünür").font = Font(
    name="Arial", bold=True, size=11, color=BORDO)
r += 1
ornek_baslik = ["Yer kodu", "Yazar", "Başlık", "Yıl", "Nüsha", "Kategori", "Kural", "Durum", "Not"]
ornek_deger = ["G-A03-007", "Boratav, Pertev Naili", "Halk Hikâyeleri ve Halk Hikâyeciliği",
               1946, 1, "Gidecek", "Y4", "Yıpranmış", "Kapağı gevşek, ciltlenmeli"]
for i, (bas, deg) in enumerate(zip(ornek_baslik, ornek_deger), start=1):
    hb = b.cell(r, i, bas)
    hb.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    hb.fill = PatternFill("solid", fgColor=BORDO)
    hd = b.cell(r + 1, i, deg)
    hd.font = Font(name="Arial", size=10)
    hd.fill = PatternFill("solid", fgColor=KREM)

b.column_dimensions["A"].width = 46
b.column_dimensions["B"].width = 30
b.column_dimensions["C"].width = 30
for s in "DEFGHI":
    b.column_dimensions[s].width = 16

# ═══════════════ 2) KURALLAR ═══════════════
k = wb.create_sheet("Kurallar")
k.append(["Kod", "Kategori", "Açıklama"])
kurallar = [
    ("Y1", "Gidecek", "Tarih Vakfı yayınları ve vakıf tarihine ait her şey"),
    ("Y2", "Gidecek", "Türkiye ekonomik/toplumsal tarihi, kent tarihi, sözlü tarih, bellek"),
    ("Y3", "Gidecek", "İmzalı, ithaflı veya ex-libris'li nüsha"),
    ("Y4", "Gidecek", "1950 öncesi baskı; Osmanlıca / eski harfli eser"),
    ("Y5", "Gidecek", "Rapor, tez, bülten, katalog, broşür — başka yerde bulunmayan"),
    ("Y6", "Gidecek", "Süreli yayının tam serisi"),
    ("Y7", "Gidecek", "Şartlı bağış koleksiyonunun parçası"),
    ("S1", "Gitse de olur", "Alanla ilgili ama ikinci veya üçüncü kopya"),
    ("S2", "Gitse de olur", "Genel başvuru kaynağı (sözlük, ansiklopedi, el kitabı)"),
    ("S3", "Gitse de olur", "Alanla dolaylı ilgili yabancı dilde kitap"),
    ("S4", "Gitse de olur", "İlgili ama güncelliğini kısmen yitirmiş"),
    ("S5", "Gitse de olur", "Sağlam ve kullanışlı, ama önceliği düşük"),
    ("K1", "Gitmeyecek", "Alan dışı: ders kitabı, sınav kitabı, popüler kurgu"),
    ("K2", "Gitmeyecek", "Güncelliğini yitirmiş mevzuat, eski yıllık/katalog fazlası"),
    ("K3", "Gitmeyecek", "Üçten fazla mükerrer nüshanın fazlası (3 nüsha kalsın)"),
    ("K4", "Gitmeyecek", "Ağır hasarlı ve kolay bulunabilir"),
    ("K5", "Gitmeyecek", "Süreli yayının dağınık tek sayısı (seri değil)"),
    ("K6", "Gitmeyecek", "Tanıtım/promosyon basılı malzeme"),
    ("M1", "Belirsiz", "On saniyede karar verilemedi"),
    ("M2", "Belirsiz", "Değerli olabilir şüphesi var"),
    ("M3", "Belirsiz", "Yabancı dilde, içerik anlaşılmadı"),
    ("M4", "Belirsiz", "Kurallar birbiriyle çelişiyor"),
]
for satir in kurallar:
    k.append(list(satir))
for c in range(1, 4):
    h = k.cell(1, c)
    h.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=BORDO)
for row in range(2, len(kurallar) + 2):
    for c in range(1, 4):
        k.cell(row, c).font = Font(name="Arial", size=10)
k.column_dimensions["A"].width = 9
k.column_dimensions["B"].width = 17
k.column_dimensions["C"].width = 66
k.freeze_panes = "A2"

# ═══════════════ 3) ENVANTER ═══════════════
e = wb.create_sheet("Envanter")
# Yer kodu: G-A03-007  →  mekân · raf · sıra · o sıradaki kaçıncı kitap
basliklar = ["Kayıt no", "Yer kodu", "Mekân", "Raf", "Sıra", "Sıra no",
             "Yazar", "Başlık", "Yıl", "Nüsha",
             "Kategori", "Kural", "Durum", "Not", "Kaydeden", "Kutu no", "Tarih"]
e.append(basliklar)
for c in range(1, len(basliklar) + 1):
    h = e.cell(1, c)
    h.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=BORDO)
    h.alignment = Alignment(vertical="center")
e.row_dimensions[1].height = 24

# Kayıt no: satır dolduğunda kendiliğinden numaralanır
for row in range(2, SATIR + 1):
    e.cell(row, 1, f'=IF(H{row}="","",ROW()-1)').font = Font(name="Arial", size=10, color="74686E")
    for c in range(2, len(basliklar) + 1):
        e.cell(row, c).font = Font(name="Arial", size=10)

genislik = [9, 14, 8, 6, 6, 8, 26, 42, 7, 8, 15, 8, 14, 30, 14, 10, 12]
for i, g in enumerate(genislik, start=1):
    e.column_dimensions[get_column_letter(i)].width = g
e.freeze_panes = "C2"
e.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{SATIR}"

# Açılır listeler
dv_kat = DataValidation(type="list", formula1='"Gidecek,Gitse de olur,Gitmeyecek,Belirsiz"',
                        allow_blank=True, showDropDown=False)
dv_kat.error = "Listeden bir kategori seçin."
dv_kat.errorTitle = "Geçersiz kategori"
e.add_data_validation(dv_kat)
dv_kat.add(f"K2:K{SATIR}")

dv_kural = DataValidation(type="list", formula1=f"=Kurallar!$A$2:$A${len(kurallar)+1}",
                          allow_blank=True, showDropDown=False)
e.add_data_validation(dv_kural)
dv_kural.add(f"L2:L{SATIR}")

dv_durum = DataValidation(type="list", formula1='"Sağlam,Yıpranmış,Küflü/böcekli"',
                          allow_blank=True, showDropDown=False)
e.add_data_validation(dv_durum)
dv_durum.add(f"M2:M{SATIR}")

# Kategoriye göre renk
renkler = [("Gidecek", "E8F3EA"), ("Gitse de olur", "FBF3E2"),
           ("Gitmeyecek", "FBEAEC"), ("Belirsiz", "E9F1F7")]
for deger, renk in renkler:
    e.conditional_formatting.add(
        f"K2:K{SATIR}",
        CellIsRule(operator="equal", formula=[f'"{deger}"'], fill=PatternFill("solid", bgColor=renk)))
# Küflü/böcekli satırı belirgin olsun
e.conditional_formatting.add(
    f"M2:M{SATIR}",
    CellIsRule(operator="equal", formula=['"Küflü/böcekli"'],
               fill=PatternFill("solid", bgColor="F6CED3"), font=Font(bold=True, color="8A1230")))

# ═══════════════ 4) KUTULAR ═══════════════
ku = wb.create_sheet("Kutular")
ku_bas = ["Kutu no", "Kaynak sıra", "Yer kodu aralığı", "Hedef bölüm", "Paketleyen",
          "Tarih", "İçindeki kitap"]
ku.append(ku_bas)
for c in range(1, len(ku_bas) + 1):
    h = ku.cell(1, c)
    h.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=BORDO)
for row in range(2, 301):
    ku.cell(row, 7, f'=IF(A{row}="","",COUNTIF(Envanter!$P$2:$P${SATIR},A{row}))')
    for c in range(1, 8):
        ku.cell(row, c).font = Font(name="Arial", size=10)
for i, g in enumerate([11, 14, 20, 24, 18, 12, 15], start=1):
    ku.column_dimensions[get_column_letter(i)].width = g
ku.freeze_panes = "A2"

# ═══════════════ 5) ÖZET ═══════════════
o = wb.create_sheet("Özet", 1)
o["A1"] = "Özet — kendiliğinden hesaplanır"
o["A1"].font = Font(name="Arial", bold=True, size=15, color=BORDO)
o["A2"] = "Bu sayfaya elle bir şey yazmayın."
o["A2"].font = Font(name="Arial", size=9.5, color="74686E")

E = f"Envanter!$K$2:$K${SATIR}"      # Kategori
D = f"Envanter!$M$2:$M${SATIR}"      # Durum
K = f"Envanter!$L$2:$L${SATIR}"      # Kural

satirlar = [
    ("İLERLEME", None, None),
    ("Kaydedilen kitap", f'=COUNTIF(Envanter!$H$2:$H${SATIR},"<>")', "#,##0"),
    ("Tahmini toplam (Başla sayfası)", f"='Başla'!$B${ilk_hesap}", "#,##0"),
    ("Tamamlanan oran", f"=IF('Başla'!$B${ilk_hesap}=0,0,COUNTIF(Envanter!$H$2:$H${SATIR},\"<>\")/'Başla'!$B${ilk_hesap})", "0%"),
    ("", None, None),
    ("KARAR DAĞILIMI", None, None),
    ("Gidecek", f'=COUNTIF({E},"Gidecek")', "#,##0"),
    ("Gitse de olur", f'=COUNTIF({E},"Gitse de olur")', "#,##0"),
    ("Gitmeyecek", f'=COUNTIF({E},"Gitmeyecek")', "#,##0"),
    ("Belirsiz", f'=COUNTIF({E},"Belirsiz")', "#,##0"),
    ("Belirsiz oranı  (%15'i geçerse kurallar yetersiz)",
     f'=IF(COUNTIF({E},"<>")=0,0,COUNTIF({E},"Belirsiz")/COUNTIF({E},"<>"))', "0%"),
    ("", None, None),
    ("HEDEFE GÖRE DURUM", None, None),
    ("Yeni binaya sığacak (Başla sayfası)", f"='Başla'!$B${ilk_hesap+1}", "#,##0"),
    ("Gidecek + Gitse de olur", f'=COUNTIF({E},"Gidecek")+COUNTIF({E},"Gitse de olur")', "#,##0"),
    ("Yalnızca Gidecek", f'=COUNTIF({E},"Gidecek")', "#,##0"),
    ("", None, None),
    ("FİZİKSEL DURUM", None, None),
    ("Sağlam", f'=COUNTIF({D},"Sağlam")', "#,##0"),
    ("Yıpranmış", f'=COUNTIF({D},"Yıpranmış")', "#,##0"),
    ("Küflü/böcekli  — ayrı alanda mı?", f'=COUNTIF({D},"Küflü/böcekli")', "#,##0"),
    ("", None, None),
    ("KUTULAMA", None, None),
    ("Açılan kutu", '=COUNTIF(Kutular!$A$2:$A$300,"<>")', "#,##0"),
    ("Kutulanan kitap", f'=COUNTIF(Envanter!$P$2:$P${SATIR},"<>")', "#,##0"),
]
r = 4
for ad, form, bicim in satirlar:
    if ad == "" and form is None:
        r += 1
        continue
    h = o.cell(r, 1, ad)
    if form is None:
        h.font = Font(name="Arial", bold=True, size=10.5, color=BORDO)
        o.cell(r, 1).fill = PatternFill("solid", fgColor=ACIK)
        o.cell(r, 2).fill = PatternFill("solid", fgColor=ACIK)
    else:
        h.font = Font(name="Arial", size=10.5)
        d = o.cell(r, 2, form)
        d.font = Font(name="Arial", bold=True, size=11)
        d.number_format = bicim
        d.alignment = Alignment(horizontal="right")
    r += 1

# Kural kodu dağılımı
r += 1
o.cell(r, 1, "KURAL KODU DAĞILIMI").font = Font(name="Arial", bold=True, size=10.5, color=BORDO)
o.cell(r, 1).fill = PatternFill("solid", fgColor=ACIK)
o.cell(r, 2).fill = PatternFill("solid", fgColor=ACIK)
r += 1
for i, (kod, kat, aciklama) in enumerate(kurallar):
    o.cell(r, 1, f"{kod} — {aciklama}").font = Font(name="Arial", size=9.5, color="57474F")
    d = o.cell(r, 2, f'=COUNTIF({K},"{kod}")')
    d.font = Font(name="Arial", size=10)
    d.number_format = "#,##0"
    d.alignment = Alignment(horizontal="right")
    r += 1

o.column_dimensions["A"].width = 52
o.column_dimensions["B"].width = 14

wb.save("/home/claude/kutuphane/kitap-envanteri.xlsx")
print("yazıldı: kitap-envanteri.xlsx")
