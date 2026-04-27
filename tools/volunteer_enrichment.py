#!/usr/bin/env python3
"""Build a volunteer-enrichment preview from the PNB paydaş + zaman çizelgesi workbooks.

The script never writes to Firestore. It produces three JSON files that a
coordinator/admin reviews before applying patches via the existing tooling:

- enrichment_preview_updates.json   — patches for users matched by fullName
- enrichment_preview_prereg.json    — names with no matching user, ready for
                                       the preregistered collection
- enrichment_preview_ambiguous.json — names that match more than one user

Matching is exact after a Turkish-aware normalization (İŞĞÜÖÇ → ISGUOC then
lowercased and whitespace-collapsed). The Firestore field on users docs is
``fullName`` — Firebase Auth's ``displayName`` is mirrored into ``fullName`` at
first sign-in (see auth/auth.js).

Patches follow a fill-the-blank rule: existing non-empty values on the user
doc are never overwritten. New volunteer-profile fields (specialty,
availabilityDays, projectExpectation) are always added.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SPECIALTY_BY_NORMALIZED_LABEL: dict[str, str] = {
    "dijitallestirme": "dijitallestirme",
    "osmanlica": "osmanlica",
    "teknik destek": "teknik_destek",
    "web sitesi alt yapisi": "web_altyapi",
    "gonullu koordinasyonu": "gonullu_koordinasyonu",
    "arsivcilik": "arsivcilik",
    "dokumantasyon sistemi": "dokumantasyon_sistemi",
    "gorsel isitsel envanter": "gorsel_isitsel_envanter",
    "mimari projeler envanteri": "mimari_projeler_envanteri",
    "dijital envanter": "dijital_envanter",
    "gecmis envanterlerin ayiklanmasi": "gecmis_envanter_ayiklama",
}


DAY_PREFIX_BY_NORMALIZED: dict[str, str] = {
    "pztesi": "mon",
    "pazartesi": "mon",
    "sali": "tue",
    "cars": "wed",
    "carsamba": "wed",
    "pers": "thu",
    "persembe": "thu",
    "cuma": "fri",
}


# Fields that already exist on user docs and must NOT be overwritten when
# non-empty. Listed alongside the source key in our enrichment record so
# `merge_patch` can compare values consistently.
FILL_BLANK_EXISTING_FIELDS = ("phone", "department", "city", "profession", "university", "projectExpectation")


def normalize_text(value: Any) -> str:
    """Turkish-aware normalization: İŞĞÜÖÇ → ISGUOC, lowercased, whitespace collapsed."""
    text = "" if value is None else str(value)
    replacements = {
        "ı": "i", "İ": "I",
        "ş": "s", "Ş": "S",
        "ğ": "g", "Ğ": "G",
        "ü": "u", "Ü": "U",
        "ö": "o", "Ö": "O",
        "ç": "c", "Ç": "C",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def digits_only(value: Any) -> str:
    if value in (None, ""):
        return ""
    return re.sub(r"\D", "", str(value))


def derive_specialty_code(value: Any) -> str | None:
    normalized = normalize_text(value)
    if not normalized:
        return None
    return SPECIALTY_BY_NORMALIZED_LABEL.get(normalized)


def derive_slot_code(header: str) -> str | None:
    """Map a column header like 'Pztesi-1' or 'Çarş-2' to 'mon-am' / 'wed-pm'."""
    normalized = normalize_text(header)
    if not normalized:
        return None
    # Find the day prefix as the longest matching key.
    match_key = None
    for key in sorted(DAY_PREFIX_BY_NORMALIZED.keys(), key=len, reverse=True):
        if normalized.startswith(key):
            match_key = key
            break
    if not match_key:
        return None
    day_code = DAY_PREFIX_BY_NORMALIZED[match_key]
    slot_code = "pm" if "2" in normalized else "am"
    return f"{day_code}-{slot_code}"


def derive_city(value: Any) -> tuple[str | None, str | None]:
    """Return (canonical_city, warning) — canonical is 'istanbul'/'ankara' or None."""
    if value in (None, ""):
        return None, None
    normalized = normalize_text(value)
    if normalized == "istanbul":
        return "istanbul", None
    if normalized == "ankara":
        return "ankara", None
    return None, f"unrecognized city: {clean_string(value)!r}"


def load_users(users_path: Path) -> list[dict[str, Any]]:
    raw = json.loads(users_path.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        return [{"id": doc_id, **data} for doc_id, data in raw.items()]
    return list(raw)


def index_users_by_normalized_name(users: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for user in users:
        name = user.get("fullName") or ""
        key = normalize_text(name)
        if key:
            by_name[key].append(user)
    return by_name


def parse_zaman_cizelgesi(path: Path) -> dict[str, dict[str, Any]]:
    """Return {normalized_name: {fullName, specialty[], availabilityDays[], sourceRow}}."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sayfa1"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [clean_string(value) for value in rows[0]]
    # Slot columns: any header that maps to a known day slot.
    slot_columns: list[tuple[int, str]] = []
    for col_index, header in enumerate(headers):
        slot = derive_slot_code(header)
        if slot:
            slot_columns.append((col_index, slot))
    # Konu (primary subject) is column index 1; the unnamed third header
    # (column index 2) is the secondary subject column.
    primary_col, secondary_col = 1, 2
    by_name: dict[str, dict[str, Any]] = {}
    for row_index, row in enumerate(rows[1:], start=2):
        full_name = clean_string(row[0]) if row else ""
        if not full_name:
            continue
        key = normalize_text(full_name)
        specialties: list[str] = []
        for source_col in (primary_col, secondary_col):
            code = derive_specialty_code(row[source_col]) if source_col < len(row) else None
            if code and code not in specialties:
                specialties.append(code)
        availability: list[str] = []
        for col_index, slot_code in slot_columns:
            cell = row[col_index] if col_index < len(row) else None
            if normalize_text(cell) == "x" and slot_code not in availability:
                availability.append(slot_code)
        by_name[key] = {
            "fullName": full_name,
            "normalizedName": key,
            "specialty": specialties,
            "availabilityDays": availability,
            "sourceRow": row_index,
        }
    return by_name


def parse_paydas(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sayfa1"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [normalize_text(value) for value in rows[0]]

    def col_index(*candidates: str) -> int | None:
        for candidate in candidates:
            target = normalize_text(candidate)
            for index, header in enumerate(headers):
                if header == target:
                    return index
        return None

    name_col = col_index("paydaş")
    email_col = col_index("email")
    phone_col = col_index("telefon")
    profession_col = col_index("meslek")
    university_col = col_index("universite", "üniversite")
    department_col = col_index("bölüm", "bolum")
    city_col = col_index("şehir", "sehir")
    expectation_col = col_index("projeden beklentisi")

    by_name: dict[str, dict[str, Any]] = {}
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(clean_string(value) for value in row):
            continue
        full_name = clean_string(row[name_col]) if name_col is not None else ""
        if not full_name:
            continue
        key = normalize_text(full_name)
        city, city_warning = derive_city(row[city_col]) if city_col is not None else (None, None)
        record = {
            "fullName": full_name,
            "normalizedName": key,
            "email": clean_string(row[email_col]).lower() if email_col is not None else "",
            "phone": digits_only(row[phone_col]) if phone_col is not None else "",
            "profession": clean_string(row[profession_col]) if profession_col is not None else "",
            "university": clean_string(row[university_col]) if university_col is not None else "",
            "department": clean_string(row[department_col]) if department_col is not None else "",
            "city": city,
            "cityRaw": clean_string(row[city_col]) if city_col is not None else "",
            "cityWarning": city_warning,
            "projectExpectation": clean_string(row[expectation_col]) if expectation_col is not None else "",
            "sourceRow": row_index,
        }
        by_name[key] = record
    return by_name


def merge_records(
    paydas: dict[str, dict[str, Any]],
    schedule: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Union the two name-keyed maps; one record per distinct normalized name."""
    merged: dict[str, dict[str, Any]] = {}
    for key, record in paydas.items():
        merged[key] = {
            "fullName": record["fullName"],
            "normalizedName": key,
            "sources": {"paydas": record["sourceRow"]},
            "phone": record["phone"],
            "profession": record["profession"],
            "university": record["university"],
            "department": record["department"],
            "city": record["city"],
            "cityRaw": record["cityRaw"],
            "cityWarning": record["cityWarning"],
            "projectExpectation": record["projectExpectation"],
            "email": record["email"],
            "specialty": [],
            "availabilityDays": [],
        }
    for key, record in schedule.items():
        if key in merged:
            slot = merged[key]
            slot["sources"]["schedule"] = record["sourceRow"]
            slot["specialty"] = record["specialty"]
            slot["availabilityDays"] = record["availabilityDays"]
        else:
            merged[key] = {
                "fullName": record["fullName"],
                "normalizedName": key,
                "sources": {"schedule": record["sourceRow"]},
                "phone": "",
                "profession": "",
                "university": "",
                "department": "",
                "city": None,
                "cityRaw": "",
                "cityWarning": None,
                "projectExpectation": "",
                "email": "",
                "specialty": record["specialty"],
                "availabilityDays": record["availabilityDays"],
            }
    return sorted(merged.values(), key=lambda item: item["normalizedName"])


def build_patch(record: dict[str, Any], existing_user: dict[str, Any]) -> dict[str, Any]:
    """Produce a sparse patch: only fields that are blank on the existing user
    or new volunteer-profile fields. Returns {} if nothing to write."""
    patch: dict[str, Any] = {}

    def existing_blank(field: str) -> bool:
        value = existing_user.get(field)
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    if record["phone"] and existing_blank("phone"):
        patch["phone"] = record["phone"]
    if record["profession"] and existing_blank("profession"):
        patch["profession"] = record["profession"]
    if record["university"] and existing_blank("university"):
        patch["university"] = record["university"]
    if record["department"] and existing_blank("department"):
        patch["department"] = record["department"]
    if record["city"] and existing_blank("city"):
        patch["city"] = record["city"]
    if record["projectExpectation"] and existing_blank("projectExpectation"):
        patch["projectExpectation"] = record["projectExpectation"]
    # Volunteer-profile fields: always overwrite-with-union? The spec says
    # additive-only. Existing user docs do not yet carry these fields, so on
    # first import they are simply set. If a future re-run encounters an
    # existing array we union-by-set to avoid clobbering manual edits.
    if record["specialty"]:
        existing = existing_user.get("specialty") or []
        merged = list(dict.fromkeys([*existing, *record["specialty"]]))
        if merged != existing:
            patch["specialty"] = merged
    if record["availabilityDays"]:
        existing = existing_user.get("availabilityDays") or []
        merged = list(dict.fromkeys([*existing, *record["availabilityDays"]]))
        if merged != existing:
            patch["availabilityDays"] = merged
    return patch


def classify(
    merged_records: list[dict[str, Any]],
    users_by_name: dict[str, list[dict[str, Any]]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    updates: list[dict[str, Any]] = []
    prereg: list[dict[str, Any]] = []
    ambiguous: list[dict[str, Any]] = []
    for record in merged_records:
        candidates = users_by_name.get(record["normalizedName"], [])
        review_flags: list[str] = []
        if record.get("cityWarning"):
            review_flags.append(record["cityWarning"])
        if len(candidates) == 1:
            user = candidates[0]
            patch = build_patch(record, user)
            updates.append({
                "fullName": record["fullName"],
                "normalizedName": record["normalizedName"],
                "userId": user.get("id") or user.get("uid"),
                "userEmail": user.get("email"),
                "sources": record["sources"],
                "patch": patch,
                "patchEmpty": not patch,
                "reviewFlags": review_flags,
            })
        elif len(candidates) >= 2:
            ambiguous.append({
                "fullName": record["fullName"],
                "normalizedName": record["normalizedName"],
                "sources": record["sources"],
                "candidates": [
                    {
                        "userId": user.get("id") or user.get("uid"),
                        "fullName": user.get("fullName"),
                        "email": user.get("email"),
                        "department": user.get("department"),
                        "role": user.get("role"),
                    }
                    for user in candidates
                ],
                "reviewFlags": review_flags,
            })
        else:
            prereg.append({
                "fullName": record["fullName"],
                "normalizedName": record["normalizedName"],
                "sources": record["sources"],
                "email": record["email"],
                "phone": record["phone"],
                "profession": record["profession"],
                "university": record["university"],
                "department": record["department"],
                "city": record["city"],
                "projectExpectation": record["projectExpectation"],
                "specialty": record["specialty"],
                "availabilityDays": record["availabilityDays"],
                "reviewFlags": review_flags,
            })
    return updates, prereg, ambiguous


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--paydas", type=Path, required=True, help="Path to PNB Paydaş workbook.")
    parser.add_argument("--schedule", type=Path, required=True, help="Path to PNB Gönüllü zaman çizelgesi workbook.")
    parser.add_argument("--users-json", type=Path, default=Path("backups/2026-04-17/users.json"), help="JSON export of the users collection used for matching.")
    parser.add_argument("--out-dir", type=Path, default=Path("tools"), help="Directory where preview JSON files are written.")
    args = parser.parse_args()

    users = load_users(args.users_json.resolve())
    users_by_name = index_users_by_normalized_name(users)

    paydas = parse_paydas(args.paydas.resolve())
    schedule = parse_zaman_cizelgesi(args.schedule.resolve())
    merged = merge_records(paydas, schedule)
    updates, prereg, ambiguous = classify(merged, users_by_name)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    (args.out_dir / "enrichment_preview_updates.json").write_text(
        json.dumps({"count": len(updates), "items": updates}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (args.out_dir / "enrichment_preview_prereg.json").write_text(
        json.dumps({"count": len(prereg), "items": prereg}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (args.out_dir / "enrichment_preview_ambiguous.json").write_text(
        json.dumps({"count": len(ambiguous), "items": ambiguous}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"users loaded: {len(users)}")
    print(f"paydaş rows: {len(paydas)}")
    print(f"schedule rows: {len(schedule)}")
    print(f"merged distinct names: {len(merged)}")
    print(f"updates: {len(updates)}  (empty patches: {sum(1 for item in updates if item['patchEmpty'])})")
    print(f"prereg: {len(prereg)}")
    print(f"ambiguous: {len(ambiguous)}")


if __name__ == "__main__":
    main()
