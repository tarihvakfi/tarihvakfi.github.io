#!/usr/bin/env python3
"""Audit and build the public Sayim Defteri dashboard aggregate from Excel."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "_audit" / "input"
OUTPUT_DIR = ROOT / "_audit" / "output"
REPORT_PATH = OUTPUT_DIR / "public_dashboard_audit.md"
SUMMARY_JSON_PATH = OUTPUT_DIR / "public_summary_from_excel.json"
SNAPSHOT_PATH = ROOT / "js" / "snapshot.js"
PROJECT_ID = "pnb"
PUBLIC_TZ = timezone(timedelta(hours=3))

TR_WEEKDAYS = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
TR_MONTHS = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
UNNAMED = "Adı belirtilmeyen gönüllü"
HIDDEN = "İsmini gizlemeyi tercih eden gönüllü"
PUBLIC_ROLE_BY_NAME = {
    "gulistan eren": "Gönüllü Koordinatörü",
}
TRACK_ORDER = [
    "tarama",
    "envanter",
    "kurumsal_bellek",
    "proje_basvuru",
    "osmanlica",
    "egitim",
    "koordinasyon",
    "ars_web",
    "kodlama_kontrol",
    "diger",
]
TRACK_LABELS = {
    "tarama": "Tarama (belge & kartpostal)",
    "envanter": "Envanter (afiş, görsel-işitsel)",
    "kurumsal_bellek": "Kurum belleği (Karar Def., G.K.)",
    "proje_basvuru": "Proje çalışmaları",
    "osmanlica": "Osmanlıca çeviri",
    "egitim": "Eğitim (İş Bankası Müzesi)",
    "koordinasyon": "Koordinasyon & planlama",
    "ars_web": "Arşiv-web & IT",
    "kodlama_kontrol": "Kodlama & kontrol",
    "diger": "Diğer çalışma",
}


def ascii_fold(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("ı", "i").replace("İ", "I")
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


TR_MONTH_LOOKUP = {ascii_fold(name).lower(): idx + 1 for idx, name in enumerate(TR_MONTHS)}
TR_MONTH_LOOKUP["harizan"] = 6


def slugify(value: Any) -> str:
    folded = ascii_fold(value).lower()
    folded = re.sub(r"[^a-z0-9]+", "_", folded)
    return folded.strip("_")


def header_key(value: Any) -> str:
    return slugify(value)


def workbook_path() -> Path:
    candidates = sorted(INPUT_DIR.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No .xlsx file found under {INPUT_DIR}")
    return candidates[0]


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_locale_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        return number if math.isfinite(number) else 0.0
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_done_total(value: Any) -> tuple[float, float]:
    if value is None or value == "":
        return 0.0, 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 0.0, float(value)
    text = str(value).strip()
    match = re.match(r"^([\d.,]+)\s*/\s*([\d.,]+)$", text)
    if match:
        return parse_locale_number(match.group(1)), parse_locale_number(match.group(2))
    return 0.0, parse_locale_number(text)


def parse_sheet_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value).strip()
    if not text:
        return None
    match = re.match(r"^(\d{1,2})\s+([A-Za-zÇĞİÖŞÜçğıöşü]+)(?:\s+(\d{4}))?$", text)
    if match:
        month = TR_MONTH_LOOKUP.get(ascii_fold(match.group(2)).lower())
        if month:
            year = int(match.group(3)) if match.group(3) else datetime.now(tz=PUBLIC_TZ).year
            return date(year, month, int(match.group(1)))
    match = re.match(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})$", text)
    if match:
        year = int(match.group(3))
        if year < 100:
            year += 2000
        return date(year, int(match.group(2)), int(match.group(1)))
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def parse_sheet_datetime(value: Any) -> datetime | None:
    d = parse_sheet_date(value)
    return datetime.combine(d, time(12, 0), tzinfo=PUBLIC_TZ) if d else None


def format_day_month(value: date) -> str:
    return f"{value.day} {TR_MONTHS[value.month - 1]}"


def format_period_label(start: date, today: date, full_end: date) -> str:
    return f"{format_date_range_label(start, full_end)} haftası · bugüne kadar"


def format_date_range_label(start: date, end: date) -> str:
    if start.month == end.month:
        return f"{start.day}–{end.day} {TR_MONTHS[start.month - 1]}"
    return f"{start.day} {TR_MONTHS[start.month - 1]} – {end.day} {TR_MONTHS[end.month - 1]}"


def public_box_label(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        if value.year == 2026 and value.month <= 12 and value.day <= 31:
            return f"{value.month}-{value.day}"
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        if value.year == 2026 and value.month <= 12 and value.day <= 31:
            return f"{value.month}-{value.day}"
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if float(value).is_integer() else str(value).replace(".", ",")
    return str(value).strip()


def normalize_box(value: Any) -> str:
    label = ascii_fold(public_box_label(value)).lower().strip()
    match = re.fullmatch(r"(\d+)\s*-\s*(\d+)", label)
    if match:
        return f"range_{match.group(1)}_{match.group(2)}"
    return re.sub(r"[^a-z0-9]+", "", label)


def is_unsafe_public_identifier(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", text):
        return True
    if re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", text):
        return True
    if re.fullmatch(r"[0-9a-fA-F]{12,}", text):
        return True
    if re.match(r"^(sheet|row|uid|user|firebase|google|apps?|script|token|id)[_:-]", text, re.I):
        return True
    compact = re.sub(r"[^A-Za-z0-9]", "", text)
    if len(compact) >= 24:
        return True
    if len(compact) >= 16 and re.search(r"\d", compact) and re.search(r"[A-Za-z]", compact):
        words = re.findall(r"[A-Za-zÇĞİÖŞÜçğıöşü]{2,}", text)
        vowels = re.findall(r"[aeıioöuüAEIİOÖUÜ]", text)
        if len(words) < 2 or len(vowels) < 2:
            return True
    return False


def safe_public_text(value: Any, max_len: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", text):
        return ""
    if is_unsafe_public_identifier(text):
        return ""
    return text if len(text) <= max_len else text[: max_len - 3].strip() + "..."


def safe_public_long_text(value: Any, max_len: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", text):
        return ""
    return text if len(text) <= max_len else text[: max_len - 3].strip() + "..."


def public_work_title(row: dict[str, Any]) -> str:
    return safe_public_text(
        row.get("calisma_alani")
        or row.get("calisma")
        or row.get("is_alani")
        or row.get("is_tanimi")
        or row.get("devam_eden_calisma")
        or row.get("notlar")
        or row.get("fon_adi")
        or row.get("fon")
        or "",
        140,
    )


def public_work_detail(row: dict[str, Any]) -> str:
    return safe_public_long_text(
        row.get("devam_eden_calisma")
        or row.get("notlar")
        or row.get("yapilan_calismaya_iliskin_sayisal_bilgi")
        or row.get("devam")
        or row.get("yapilan_is")
        or row.get("aciklama")
        or "",
        260,
    )


def normalize_volunteer_name(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    text = re.sub(r"^[\s:;,\-–—]+|[\s:;,\-–—]+$", "", text)
    if not text or is_unsafe_public_identifier(text):
        return ""
    if not re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", text):
        return ""
    if len(text) > 64:
        return ""
    return text


def explicit_opt_out(row: dict[str, Any]) -> bool:
    keys = {
        "public_credit",
        "credit_visible",
        "hide_name",
        "name_hidden",
        "publiccredit",
        "creditvisible",
        "hidename",
        "ad_gizli",
        "ismini_gizle",
    }
    for key, value in row.items():
        if key not in keys:
            continue
        text = ascii_fold(value).strip().lower()
        if key in {"credit_visible", "creditvisible"}:
            if text in {"false", "0", "no", "hayir", "hayır"}:
                return True
        elif key in {"public_credit", "publiccredit"}:
            if text in {"false", "0", "no", "hayir", "hayır"}:
                return True
        elif text in {"true", "1", "yes", "evet", "gizli", "hide", "hidden"}:
            return True
    return False


def get_volunteer_display_name(row: dict[str, Any]) -> tuple[str, str]:
    if explicit_opt_out(row):
        return HIDDEN, "opt_out"
    explicit_keys = ("public_display_name", "publicdisplayname", "kamusal_ad", "kamusalad")
    for key in explicit_keys:
        name = normalize_volunteer_name(row.get(key))
        if name:
            return name, "explicit_public_display"
    name_keys = (
        "paydas",
        "paydaş",
        "kaydi_olusuran",
        "kaydi_olusturan",
        "kaydi_olusturan_2",
        "kaydi_olusuran_2",
        "_sheet_person",
        "volunteer_name",
        "ad_soyad",
    )
    for key in name_keys:
        raw = row.get(key)
        if raw and is_unsafe_public_identifier(raw):
            return UNNAMED, "unsafe_identifier"
        name = normalize_volunteer_name(raw)
        if name:
            return name, "name"
    first = normalize_volunteer_name(row.get("first_name") or row.get("ad") or row.get("isim"))
    last = normalize_volunteer_name(row.get("last_name") or row.get("soyad") or row.get("soyisim"))
    combined = normalize_volunteer_name(f"{first} {last}".strip())
    if combined:
        return combined, "first_last"
    if first:
        return first, "first_name"
    return UNNAMED, "missing"


def private_contributor_key(label: str, row: dict[str, Any]) -> str:
    if label not in {UNNAMED, HIDDEN}:
        return ascii_fold(label).lower().strip()
    if label == UNNAMED:
        return "unnamed"
    if label == HIDDEN:
        return "hidden"
    raw = row.get("paydas") or row.get("kaydi_olusuran") or row.get("kaydi_olusturan") or row.get("_sheet_person")
    if raw and not is_unsafe_public_identifier(raw):
        return ascii_fold(raw).lower().strip()
    return f"unnamed:{row.get('_sheet')}:{row.get('_source_row')}"


def preferred_label(labels: list[str]) -> str:
    clean = [label for label in labels if label]
    if not clean:
        return UNNAMED
    if HIDDEN in clean:
        return HIDDEN
    if all(label == UNNAMED for label in clean):
        return UNNAMED
    counts = Counter(clean)
    human = [label for label in clean if label not in {UNNAMED, HIDDEN}]
    if not human:
        return counts.most_common(1)[0][0]
    return sorted(
        set(human),
        key=lambda label: (
            -sum(1 for ch in label if ch in "ÇĞİÖŞÜçğıöşü"),
            -counts[label],
            -len(label),
            ascii_fold(label).lower(),
        ),
    )[0]


def is_public_named_label(label: Any) -> bool:
    text = str(label or "").strip()
    return bool(text and text not in {UNNAMED, HIDDEN} and not is_unsafe_public_identifier(text))


def public_people_from_label(label: Any) -> list[dict[str, str]]:
    if not is_public_named_label(label):
        return []
    people = []
    seen: dict[str, str] = {}
    for part in re.split(r"\s*(?:,|;|\n|\s+-\s+)\s*", str(label or "")):
        name = normalize_volunteer_name(part)
        if not is_public_named_label(name):
            continue
        key = ascii_fold(name).lower().strip()
        seen[key] = preferred_label([seen.get(key, ""), name])
    for key, name in seen.items():
        people.append({"key": key, "label": name})
    return people


def normalize_public_role(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    folded = ascii_fold(text).lower()
    if folded in {"false", "0", "no", "hayir", "hayır", "yok", "none", "null"}:
        return ""
    if is_unsafe_public_identifier(text) or len(text) > 80:
        return ""
    if not re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", text):
        return ""
    return text


def get_public_role(row: dict[str, Any], label: str) -> str:
    role_keys = (
        "role",
        "gorev",
        "public_role",
        "publicrole",
        "display_role",
        "displayrole",
        "coordinator",
        "koordinator",
        "koordinatör",
    )
    for key in role_keys:
        if key not in row:
            continue
        raw = row.get(key)
        folded = ascii_fold(raw).lower().strip()
        if key in {"coordinator", "koordinator", "koordinatör"}:
            if folded in {"true", "1", "yes", "evet", "x"} or "koordinator" in folded:
                return "Gönüllü Koordinatörü"
        role = normalize_public_role(raw)
        if role:
            return role
    return PUBLIC_ROLE_BY_NAME.get(ascii_fold(label).lower().strip(), "")


def preferred_role(roles: list[str], label: str) -> str:
    clean = [normalize_public_role(role) for role in roles if normalize_public_role(role)]
    if clean:
        return Counter(clean).most_common(1)[0][0]
    return PUBLIC_ROLE_BY_NAME.get(ascii_fold(label).lower().strip(), "")


def sheet_person_from_title(title: str) -> str:
    parts = str(title or "").split()
    if not parts or parts[0].upper() != "PNB":
        return ""
    return parts[-1].replace("-", " - ")


def material_category(row: dict[str, Any]) -> str:
    haystack = ascii_fold(
        " ".join(str(row.get(key) or "") for key in (
            "calisma_alani", "devam_eden_calisma", "dijital_belge_kodu",
            "notlar", "fon", "fon_adi"
        ))
    ).lower()
    if "foto" in haystack or "gorsel" in haystack or "dia" in haystack:
        return "fotoğraflar"
    if "mektup" in haystack:
        return "mektuplar"
    if "kitap" in haystack:
        return "kitap metinleri"
    if "ders" in haystack:
        return "ders notları"
    if "envanter" in haystack:
        return "envanter"
    return "belgeler"


def row_project_id(row: dict[str, Any]) -> str:
    haystack = ascii_fold(
        " ".join(str(row.get(key) or "") for key in (
            "fon", "fon_adi", "calisma_alani", "devam_eden_calisma", "dijital_belge_kodu", "notlar"
        ))
    ).lower()
    return PROJECT_ID if "pnb" in haystack or "boratav" in haystack else "foundation"


def track_key_for_record(record: SourceRecord) -> str:
    if record.kind == "page":
        return "tarama"
    return track_key_from_text(" ".join([record.work_title, record.work_detail, record.material, record.public_role, record.project_id]))


def track_key_from_text(value: str) -> str:
    haystack = ascii_fold(value).lower()
    if re.search(r"\b(web|site|teknik|it)\b", haystack) or "arsiv web" in haystack or "arsiv-web" in haystack:
        return "ars_web"
    if "egitim" in haystack or "is bankasi" in haystack or "muze" in haystack:
        return "egitim"
    if "osmanlica" in haystack or "ceviri" in haystack:
        return "osmanlica"
    if any(token in haystack for token in ["envanter", "afis", "gorsel", "harita", "dvd", "video", "sozlu tarih", "aktarim", "tashih"]):
        return "envanter"
    if any(token in haystack for token in ["kronoloji", "karar defter", "genel kurul", "faaliyet rapor", "kurucu"]):
        return "kurumsal_bellek"
    if any(token in haystack for token in ["gerda", "basvuru", "proje butcesi", "proje form", "proje hazir", "proje deger", "culture civic", "salt", "fon"]):
        return "proje_basvuru"
    if any(token in haystack for token in ["koordinasyon", "koordinator", "planlama", "oryantasyon", "organizasyon", "toplanti", "gorusme"]):
        return "koordinasyon"
    if any(token in haystack for token in ["kodlama", "kontrol", "duzelt", "duzenleme", "adlandirma"]):
        return "kodlama_kontrol"
    if any(token in haystack for token in ["tarama", "sayisallastirma", "dijitallestirme", "dia", "pnb", "kutu"]):
        return "tarama"
    return "diger"


def classify_sheet(title: str) -> str:
    slug = slugify(title)
    if slug == "pnb_sayisallastirma":
        return "pnb_inventory"
    if slug == "gunluk_akis":
        return "activity"
    if slug == "gunluk_gonullu_akisi":
        return "schedule"
    if slug.startswith("pnb_") and "_zarf" not in slug and slug != "pnb_sayisallastirma":
        return "pnb_detail"
    return "other"


def infer_columns(headers: list[str]) -> dict[str, list[str]]:
    keys = [header_key(h) for h in headers]
    groups = {
        "dates": [],
        "volunteers": [],
        "boxes": [],
        "materials": [],
        "targets": [],
        "done": [],
    }
    for raw, key in zip(headers, keys):
        if "tarih" in key or "date" in key:
            groups["dates"].append(raw)
        if key in {"paydas", "kaydi_olusuran", "kaydi_olusturan"} or "ad_soyad" in key or "volunteer" in key:
            groups["volunteers"].append(raw)
        if "kutu" in key:
            groups["boxes"].append(raw)
        if "fon" in key or "calisma" in key or "dijital_belge" in key:
            groups["materials"].append(raw)
        if "sayfa_sayisi" in key or "hedef" in key:
            groups["targets"].append(raw)
        if "kontrol" in key or "kodlama" in key or "tarama" in key:
            groups["done"].append(raw)
    return groups


def read_rows(ws) -> tuple[list[str], list[dict[str, Any]]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_values = list(next(rows_iter))
    except StopIteration:
        return [], []
    headers = []
    seen: Counter[str] = Counter()
    for idx, value in enumerate(header_values):
        key = header_key(value) if not is_blank(value) else f"column_{idx + 1}"
        seen[key] += 1
        if seen[key] > 1:
            key = f"{key}_{seen[key]}"
        headers.append(key)
    out = []
    for row_idx, values in enumerate(rows_iter, start=2):
        values = list(values)
        if all(is_blank(v) for v in values):
            continue
        row = {"_source_row": row_idx, "_sheet": ws.title, "_sheet_slug": slugify(ws.title)}
        for idx, key in enumerate(headers):
            row[key] = values[idx] if idx < len(values) else None
        out.append(row)
    raw_headers = [str(v).strip() for v in header_values if not is_blank(v)]
    return raw_headers, out


@dataclass
class SourceRecord:
    kind: str
    source_type: str
    date: date | None
    when: datetime | None
    material: str
    project_id: str
    private_key: str
    public_label: str
    public_role: str
    credit_status: str
    box: str = ""
    page_units: int = 0
    row_count: int = 1
    work_title: str = ""
    work_detail: str = ""


@dataclass
class BoxInfo:
    box: str
    name: str
    target_pages: int = 0
    summary_done_pages: int = 0
    detail_done_pages: int = 0
    page_rows: int = 0
    files: int = 0
    documents: int = 0
    materials: Counter[str] = field(default_factory=Counter)
    contributor_keys: set[str] = field(default_factory=set)
    contributor_labels: dict[str, str] = field(default_factory=dict)
    last_activity: date | None = None

    @property
    def done_pages(self) -> int:
        return max(self.summary_done_pages, self.detail_done_pages)

    @property
    def remaining_pages(self) -> int | None:
        return max(0, self.target_pages - self.done_pages) if self.target_pages > 0 else None

    @property
    def percent(self) -> float | None:
        if self.target_pages <= 0:
            return None
        return round((min(self.done_pages, self.target_pages) / self.target_pages) * 100, 1)

    @property
    def status(self) -> str:
        if self.target_pages > 0 and self.done_pages >= self.target_pages:
            return "completed"
        if self.done_pages > 0 or self.page_rows > 0:
            return "active"
        return "inventory"


def load_workbook_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_info = []
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        headers, rows = read_rows(ws)
        rows_by_sheet[ws.title] = rows
        sheet_info.append({
            "title": ws.title,
            "classification": classify_sheet(ws.title),
            "rows": len(rows),
            "headers": headers,
            "inferred": infer_columns(headers),
        })
    return sheet_info, rows_by_sheet


def build_inventory(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, BoxInfo]:
    boxes: dict[str, BoxInfo] = {}
    for title, rows in rows_by_sheet.items():
        if classify_sheet(title) != "pnb_inventory":
            continue
        for row in rows:
            box = public_box_label(row.get("kutu") or row.get("kutu_no"))
            key = normalize_box(box)
            if not key:
                continue
            done, total = parse_done_total(row.get("sayfa_sayisi"))
            boxes[key] = BoxInfo(
                box=box,
                name=f"Kutu {box}",
                target_pages=round(total),
                summary_done_pages=round(done),
                files=round(parse_locale_number(row.get("dosya_sayisi"))),
                documents=round(parse_locale_number(row.get("belge_sayisi"))),
            )
    return boxes


def compute_inventory_totals(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, int]:
    total_pages = total_units = total_files = catalogued_boxes = 0
    for title, rows in rows_by_sheet.items():
        if classify_sheet(title) != "pnb_inventory":
            continue
        for row in rows:
            _, target = parse_done_total(row.get("sayfa_sayisi"))
            total_pages += target
            total_units += parse_locale_number(row.get("belge_sayisi"))
            total_files += parse_locale_number(row.get("dosya_sayisi"))
            box = public_box_label(row.get("kutu") or row.get("kutu_no"))
            if box and (target > 0 or parse_locale_number(row.get("belge_sayisi")) > 0 or parse_locale_number(row.get("dosya_sayisi")) > 0):
                catalogued_boxes += 1
    return {
        "totalPages": round(total_pages),
        "totalUnits": round(total_units),
        "totalFiles": round(total_files),
        "cataloguedBoxes": catalogued_boxes,
    }


def page_units_for_detail_rows(rows: list[dict[str, Any]]) -> list[int]:
    has_sayfa_sayisi = any(not is_blank(row.get("sayfa_sayisi")) for row in rows)
    units = []
    for row in rows:
        if has_sayfa_sayisi:
            done, total = parse_done_total(row.get("sayfa_sayisi"))
            units.append(round(done or total or 1))
        else:
            units.append(1)
    return units


def collect_records(rows_by_sheet: dict[str, list[dict[str, Any]]], boxes: dict[str, BoxInfo]) -> list[SourceRecord]:
    records: list[SourceRecord] = []
    for title, rows in rows_by_sheet.items():
        classification = classify_sheet(title)
        if classification == "pnb_detail":
            units = page_units_for_detail_rows(rows)
            sheet_person = sheet_person_from_title(title)
            for row, page_units in zip(rows, units):
                enriched = dict(row)
                enriched["_sheet_person"] = sheet_person
                label, status = get_volunteer_display_name(enriched)
                role = get_public_role(enriched, label)
                key = private_contributor_key(label, enriched)
                box = public_box_label(row.get("kutu") or row.get("kutu_no"))
                when = parse_sheet_datetime(row.get("tarih"))
                rec = SourceRecord(
                    kind="page",
                    source_type="page_detail",
                    date=when.date() if when else None,
                    when=when,
                    material=material_category(row),
                    project_id=PROJECT_ID,
                    private_key=key,
                    public_label=label,
                    public_role=role,
                    credit_status=status,
                    box=box,
                    page_units=max(1, page_units),
                    work_title=public_work_title(row),
                    work_detail=public_work_detail(row),
                )
                records.append(rec)
                box_key = normalize_box(box)
                if box_key:
                    info = boxes.setdefault(box_key, BoxInfo(box=box, name=f"Kutu {box}"))
                    info.detail_done_pages += rec.page_units
                    info.page_rows += 1
                    info.materials[rec.material] += 1
                    info.contributor_keys.add(key)
                    info.contributor_labels[key] = preferred_label([info.contributor_labels.get(key, ""), label])
                    if rec.date and (not info.last_activity or rec.date > info.last_activity):
                        info.last_activity = rec.date
        elif classification == "activity":
            for row in rows:
                label, status = get_volunteer_display_name(row)
                role = get_public_role(row, label)
                key = private_contributor_key(label, row)
                when = parse_sheet_datetime(row.get("tarih"))
                records.append(SourceRecord(
                    kind="activity",
                    source_type="activity",
                    date=when.date() if when else None,
                    when=when,
                    material=material_category(row),
                    project_id=row_project_id(row),
                    private_key=key,
                    public_label=label,
                    public_role=role,
                    credit_status=status,
                    page_units=0,
                    work_title=public_work_title(row),
                    work_detail=public_work_detail(row),
                ))
    return records


def selected_period(now_date: date, mode: str) -> dict[str, Any]:
    if mode == "rolling_7_days":
        start = now_date - timedelta(days=6)
        return {
            "mode": "rolling_7_days",
            "startDate": start.isoformat(),
            "endDate": now_date.isoformat(),
            "label": f"Güncel dönem · {format_date_range_label(start, now_date)}",
            "isPartial": False,
        }
    start = now_date - timedelta(days=now_date.weekday())
    full_end = start + timedelta(days=6)
    return {
        "mode": "calendar_week_to_date",
        "startDate": start.isoformat(),
        "endDate": now_date.isoformat(),
        "fullEndDate": full_end.isoformat(),
        "label": format_period_label(start, now_date, full_end),
        "isPartial": now_date < full_end,
    }


def records_in_period(records: list[SourceRecord], period: dict[str, Any]) -> list[SourceRecord]:
    start = date.fromisoformat(period["startDate"])
    end = date.fromisoformat(period["endDate"])
    return [r for r in records if r.date and start <= r.date <= end]


def build_by_track(period_records: list[SourceRecord]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for rec in period_records:
        if not is_public_named_label(rec.public_label):
            continue
        key = track_key_for_record(rec)
        group = groups.setdefault(key, {
            "key": key,
            "label": TRACK_LABELS.get(key, TRACK_LABELS["diger"]),
            "records": 0,
            "pageRows": 0,
            "activityRows": 0,
            "pagesDone": 0,
            "contributors": defaultdict(lambda: {"labels": [], "records": []}),
        })
        group["records"] += 1
        if rec.kind == "page":
            group["pageRows"] += 1
            group["pagesDone"] += rec.page_units or 1
        else:
            group["activityRows"] += 1
        for person in public_people_from_label(rec.public_label):
            bucket = group["contributors"][person["key"]]
            bucket["labels"].append(person["label"])
            bucket["records"].append(rec)

    rows = []
    ordered_keys = [key for key in TRACK_ORDER if key in groups] + sorted(key for key in groups if key not in TRACK_ORDER)
    for key in ordered_keys:
        group = groups[key]
        contributors = []
        for bucket in group["contributors"].values():
            recs = bucket["records"]
            label = preferred_label(bucket["labels"])
            if not is_public_named_label(label):
                continue
            role = preferred_role([rec.public_role for rec in recs], label)
            page_rows = [rec for rec in recs if rec.kind == "page"]
            activity_rows = [rec for rec in recs if rec.kind == "activity"]
            contributors.append({
                "label": label,
                "publicRole": role,
                "records": len(recs),
                "pageRows": len(page_rows),
                "activityRows": len(activity_rows),
                "pagesDone": sum(rec.page_units or 1 for rec in page_rows),
            })
        contributors.sort(key=lambda row: (-row["records"], row["label"]))
        if contributors:
            rows.append({
                "key": group["key"],
                "label": group["label"],
                "records": group["records"],
                "pageRows": group["pageRows"],
                "activityRows": group["activityRows"],
                "pagesDone": group["pagesDone"],
                "peopleCount": len(contributors),
                "contributors": contributors,
            })
    rows.sort(key=lambda row: (-row["records"], TRACK_ORDER.index(row["key"]) if row["key"] in TRACK_ORDER else 999, row["label"]))
    return rows


def date_range(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def material_list(counter: Counter[str]) -> list[dict[str, Any]]:
    total = sum(counter.values())
    return [
        {
            "material": material,
            "label": material[:1].upper() + material[1:],
            "count": count,
            "percent": round((count / total) * 100, 1) if total else 0,
        }
        for material, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]


def summarize_day(day: date, rows: list[SourceRecord]) -> dict[str, Any]:
    page_rows = [r for r in rows if r.kind == "page"]
    activity_rows = [r for r in rows if r.kind == "activity"]
    by_volunteer_key: dict[str, list[SourceRecord]] = defaultdict(list)
    for rec in rows:
        if rec.private_key:
            by_volunteer_key[rec.private_key].append(rec)
    contributors = []
    for recs in by_volunteer_key.values():
        label = preferred_label([r.public_label for r in recs])
        if not is_public_named_label(label):
            continue
        role = preferred_role([r.public_role for r in recs], label)
        c_page_rows = [r for r in recs if r.kind == "page"]
        c_activity_rows = [r for r in recs if r.kind == "activity"]
        contributors.append({
            "label": label,
            "publicRole": role,
            "records": len(recs),
            "pageRows": len(c_page_rows),
            "activityRows": len(c_activity_rows),
            "pagesDone": sum(r.page_units for r in c_page_rows),
            "workRows": public_work_rows_for_contributor(recs, 4),
        })
    contributors.sort(key=lambda item: ascii_fold(item["label"]).lower())
    volunteers = [item["label"] for item in contributors if not item["publicRole"]]
    coordination = [item for item in contributors if item["publicRole"]]
    boxes = sorted({public_box_label(r.box) for r in page_rows if r.box}, key=lambda s: ascii_fold(s))
    materials = Counter(r.material for r in rows if r.material)
    first = min((r.when for r in rows if r.when), default=None)
    last = max((r.when for r in rows if r.when), default=None)
    if rows:
        parts = []
        if page_rows:
            parts.append(f"{len(page_rows)} sayfa/detay satırı")
        if activity_rows:
            parts.append(f"{len(activity_rows)} faaliyet kaydı")
        if contributors:
            parts.append(f"{len(contributors)} kişi")
        if boxes:
            parts.append(f"{len(boxes)} kutu")
        sentence = "Bugün " + ", ".join(parts) + " işlendi."
    else:
        sentence = "Bugün için görünür katkı yok."
    return {
        "dateISO": day.isoformat(),
        "weekdayTR": TR_WEEKDAYS[day.weekday()],
        "dayNumber": day.day,
        "records": len(rows),
        "pageRows": len(page_rows),
        "activityRows": len(activity_rows),
        "pagesDone": sum(r.page_units for r in page_rows),
        "volunteersCount": len(contributors),
        "volunteerNames": volunteers,
        "coordination": coordination,
        "contributors": contributors,
        "boxesCount": len(boxes),
        "boxLabels": [f"Kutu {box}" for box in boxes],
        "materials": material_list(materials),
        "firstTime": first.isoformat() if first else None,
        "lastTime": last.isoformat() if last else None,
        "summarySentence": sentence,
    }


def public_work_rows_for_contributor(records: list[SourceRecord], limit: int = 4) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for rec in records:
        box_label = f"Kutu {public_box_label(rec.box)}" if rec.box else ""
        key = (rec.kind, rec.material, box_label, rec.work_title, rec.work_detail)
        row = groups.setdefault(key, {
            "dateISO": rec.date.isoformat() if rec.date else None,
            "kind": rec.kind,
            "material": rec.material,
            "boxLabel": box_label,
            "workTitle": rec.work_title,
            "workDetail": rec.work_detail,
            "records": 0,
            "pageRows": 0,
            "activityRows": 0,
            "pagesDone": 0,
        })
        row["records"] += 1
        if rec.kind == "page":
            row["pageRows"] += 1
            row["pagesDone"] += rec.page_units or 1
        else:
            row["activityRows"] += 1
    return sorted(
        groups.values(),
        key=lambda row: (-row["records"], ascii_fold(row["workTitle"]).lower(), ascii_fold(row["workDetail"]).lower()),
    )[:limit]


def build_public_summary(
    records: list[SourceRecord],
    boxes: dict[str, BoxInfo],
    inventory_totals: dict[str, int],
    generated_at: datetime,
    period_mode: str,
) -> dict[str, Any]:
    now_date = generated_at.astimezone(PUBLIC_TZ).date()
    period = selected_period(now_date, period_mode)
    period_records = records_in_period(records, period)
    page_records = [r for r in period_records if r.kind == "page"]
    activity_records = [r for r in period_records if r.kind == "activity"]
    all_page_records = [r for r in records if r.kind == "page"]

    start = date.fromisoformat(period["startDate"])
    end = date.fromisoformat(period["endDate"])
    by_day_lookup: dict[date, list[SourceRecord]] = defaultdict(list)
    for rec in period_records:
        if rec.date:
            by_day_lookup[rec.date].append(rec)
    by_day = [summarize_day(day, by_day_lookup.get(day, [])) for day in date_range(start, end)]

    by_material_counter = Counter(r.material for r in period_records if r.material)
    by_box_counter: dict[str, list[SourceRecord]] = defaultdict(list)
    for rec in page_records:
        key = normalize_box(rec.box)
        if key:
            by_box_counter[key].append(rec)

    boxes_payload = []
    for key, recs in by_box_counter.items():
        info = boxes.get(key) or BoxInfo(box=recs[0].box, name=f"Kutu {recs[0].box}")
        contributor_counts = Counter(r.private_key for r in recs if r.private_key)
        contributors = []
        for private_key, count in contributor_counts.most_common(5):
            label = info.contributor_labels.get(private_key) or preferred_label([r.public_label for r in recs if r.private_key == private_key])
            if not is_public_named_label(label):
                continue
            role = preferred_role([r.public_role for r in recs if r.private_key == private_key], label)
            page_count = sum(r.page_units for r in recs if r.private_key == private_key)
            contributors.append({"label": label, "publicRole": role, "records": count, "pageRows": count, "pagesDone": page_count})
        boxes_payload.append({
            "box": info.box,
            "boxLabel": f"Kutu {info.box}",
            "label": f"Kutu {info.box}",
            "boxTitle": info.name,
            "done": info.done_pages,
            "target": info.target_pages or None,
            "percent": info.percent,
            "remaining": info.remaining_pages,
            "records": len(recs),
            "pageRows": len(recs),
            "activityRows": 0,
            "periodRecords": len(recs),
            "periodPageRows": len(recs),
            "periodPagesDone": sum(r.page_units for r in recs),
            "materialCounts": material_list(Counter(r.material for r in recs)),
            "materials": material_list(Counter(r.material for r in recs)),
            "contributors": contributors,
            "topContributors": contributors,
            "contributorsCount": len(contributors),
            "lastActivityDate": info.last_activity.isoformat() if info.last_activity else None,
            "lastActivityLabel": format_day_month(info.last_activity) if info.last_activity else None,
            "status": info.status,
            "targetMissing": info.target_pages <= 0,
            "overTarget": info.target_pages > 0 and info.done_pages > info.target_pages,
        })
    boxes_payload.sort(key=lambda b: (-int(b["periodPagesDone"]), str(b["box"])))

    by_volunteer: dict[str, list[SourceRecord]] = defaultdict(list)
    for rec in period_records:
        if rec.private_key:
            by_volunteer[rec.private_key].append(rec)
    volunteer_payload = []
    for private_key, recs in sorted(by_volunteer.items(), key=lambda item: (-len(item[1]), ascii_fold(item[0]))):
        label = preferred_label([r.public_label for r in recs if r.public_label])
        if not is_public_named_label(label):
            continue
        role = preferred_role([r.public_role for r in recs], label)
        page_rows = [r for r in recs if r.kind == "page"]
        activity_rows = [r for r in recs if r.kind == "activity"]
        box_counts = Counter(public_box_label(r.box) for r in page_rows if r.box)
        box_breakdown = [
            {"box": box, "boxLabel": f"Kutu {box}", "records": count}
            for box, count in box_counts.most_common(4)
        ]
        volunteer_payload.append({
            "label": label,
            "publicRole": role,
            "records": len(recs),
            "pageRows": len(page_rows),
            "activityRows": len(activity_rows),
            "pagesDone": sum(r.page_units for r in page_rows),
            "topBox": f"Kutu {box_counts.most_common(1)[0][0]}" if box_counts else None,
            "boxes": [f"Kutu {box}" for box, _ in box_counts.most_common(4)],
            "boxBreakdown": box_breakdown,
        })
    track_payload = build_by_track(period_records)

    target_pages = inventory_totals.get("totalPages") or sum(info.target_pages for info in boxes.values())
    done_pages = sum(r.page_units for r in all_page_records)
    progress_percent = round((done_pages / target_pages) * 100, 1) if target_pages else 0
    inventory_boxes = [info for info in boxes.values() if info.target_pages or info.files or info.documents]
    completed_boxes = [info for info in inventory_boxes if info.status == "completed"]
    named_period = {r.private_key for r in period_records if r.private_key and is_public_named_label(r.public_label)}

    warnings = []
    if done_pages > target_pages > 0:
        warnings.append({"code": "pages_done_exceeds_target", "message": "Kaydedilen sayfa birimleri hedef toplamı aşıyor."})
    over_target_boxes = [info.box for info in inventory_boxes if info.target_pages > 0 and info.done_pages > info.target_pages]
    if over_target_boxes:
        warnings.append({"code": "box_done_exceeds_target", "message": f"{len(over_target_boxes)} kutuda işlenen sayı hedefi aşıyor.", "boxes": over_target_boxes[:8]})
    missing_targets = [info.box for info in boxes.values() if info.page_rows > 0 and info.target_pages <= 0]
    if missing_targets:
        warnings.append({"code": "missing_box_targets", "message": f"{len(missing_targets)} aktif kutuda hedef sayfa toplamı yok.", "boxes": missing_targets[:8]})
    unsafe_count = sum(1 for r in records if r.credit_status == "unsafe_identifier")
    if unsafe_count:
        warnings.append({"code": "unsafe_public_identifiers_suppressed", "message": f"{unsafe_count} katkı alanı teknik kimlik gibi göründüğü için isim olarak gösterilmedi."})
    missing_names = sum(1 for r in records if r.credit_status == "missing")
    if missing_names:
        warnings.append({"code": "missing_volunteer_names", "message": f"{missing_names} katkı satırında kullanılabilir gönüllü adı yok."})
    unknown_dates = sum(1 for r in records if not r.date)
    if unknown_dates:
        warnings.append({"code": "unknown_dates", "message": f"{unknown_dates} satır dönem/gün grafiğine atanabilecek tarih taşımıyor."})
    for day in by_day:
        if day["records"] > 0 and day["volunteersCount"] == 0:
            warnings.append({"code": "day_without_volunteer_names", "message": f"{day['dateISO']} kayıt içeriyor ama geçerli gönüllü adı yok."})
    if sum(day["records"] for day in by_day) != len(period_records):
        warnings.append({"code": "by_day_total_mismatch", "message": "Günlük toplam dönem kayıt toplamıyla eşleşmiyor."})

    busiest_day = max(by_day, key=lambda day: day["records"], default=None)
    return {
        "generatedAt": generated_at.astimezone(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source": {
            "name": "Tarih Vakfı Gönüllü Ağı",
            "projectId": PROJECT_ID,
            "recordsAreFullAggregate": True,
            "latestActivityCap": 50,
            "volunteerCredit": "credit-visible, ID-safe volunteer display",
        },
        "period": period,
        "totals": {
            "records": len(period_records),
            "pageRows": len(page_records),
            "activityRows": len(activity_records),
            "periodPagesDone": sum(r.page_units for r in page_records),
            "pagesDone": done_pages,
            "pagesTarget": target_pages,
            "progressPercent": progress_percent,
            "boxesTotal": len(inventory_boxes) or None,
            "boxesCatalogued": inventory_totals.get("cataloguedBoxes") or len(inventory_boxes),
            "boxesActive": len(by_box_counter),
            "boxesCompleted": len(completed_boxes),
            "boxesRemaining": max(0, (len(inventory_boxes) or 0) - len(completed_boxes)) if inventory_boxes else None,
            "volunteersActive": len(named_period),
            "volunteers": len(named_period),
            "materials": len(by_material_counter),
        },
        "byDay": by_day,
        "byMaterial": material_list(by_material_counter),
        "byBox": boxes_payload,
        "byVolunteer": volunteer_payload,
        "byTrack": track_payload,
        "highlights": {
            "busiestDay": busiest_day,
            "latestDate": max((r.date.isoformat() for r in period_records if r.date), default=None),
            "topMaterial": material_list(by_material_counter)[0] if by_material_counter else None,
            "firstCompletedBox": None,
        },
        "warnings": warnings,
    }


def latest_activity(records: list[SourceRecord], limit: int = 50, max_date: date | None = None) -> list[dict[str, Any]]:
    dated = [
        r for r in records
        if r.when and r.date and (max_date is None or r.date <= max_date) and is_public_named_label(r.public_label)
    ]
    dated.sort(key=lambda r: r.when or datetime.min.replace(tzinfo=PUBLIC_TZ), reverse=True)
    items: list[dict[str, Any]] = []
    for rec in dated[:limit]:
        item = {
            "when": rec.when.astimezone(timezone.utc).isoformat().replace("+00:00", "Z") if rec.when else None,
            "dateISO": rec.date.isoformat() if rec.date else None,
            "kind": rec.kind,
            "recordType": rec.source_type,
            "material": rec.material,
            "projectId": rec.project_id,
            "volunteerLabel": rec.public_label,
            "publicRole": rec.public_role,
            "boxLabel": f"Kutu {rec.box}" if rec.box else None,
            "pagesDone": rec.page_units,
        }
        if rec.work_title:
            item["workTitle"] = rec.work_title
        if rec.work_detail:
            item["workDetail"] = rec.work_detail
        items.append(item)
    return items


def build_payload(summary: dict[str, Any], records: list[SourceRecord]) -> dict[str, Any]:
    totals = summary["totals"]
    max_date = date.fromisoformat(summary["period"]["endDate"])
    return {
        "generatedAt": summary["generatedAt"],
        "publicSummary": summary,
        "latestActivity": latest_activity(records, 50, max_date),
        "content": {},
        "stats": {
            "projects": {
                PROJECT_ID: {
                    "totalPages": totals["pagesTarget"],
                    "donePages": totals["pagesDone"],
                    "cataloguedBoxes": totals["boxesCatalogued"],
                    "doneUnits": None,
                    "totalUnits": None,
                }
            }
        },
        "ticker": [],
    }


def write_snapshot(payload: dict[str, Any], path: Path) -> None:
    safe_payload = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    text = (
        "// Auto-generated public Gönüllü Emek Günlüğü snapshot.\n"
        "// Contains full aggregates plus a capped latestActivity feed; no raw workbook rows.\n"
        "window.TVF_PUBLIC_DATA = "
        + safe_payload
        + ";\n"
        "window.__SNAPSHOT__ = {ok:true,generatedAt:window.TVF_PUBLIC_DATA.generatedAt,data:window.TVF_PUBLIC_DATA};\n"
    )
    path.write_text(text, encoding="utf-8")


def extract_snapshot_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8")
    match = re.search(r"window\.TVF_PUBLIC_DATA\s*=\s*(\{.*?\})\s*;\s*window\.__SNAPSHOT__", text, re.S)
    if not match:
        match = re.search(r"window\.__SNAPSHOT__\s*=\s*(\{.*\})\s*;\s*$", text, re.S)
        if not match:
            return None
        obj = json.loads(match.group(1))
        return (obj.get("data") if obj.get("data") else obj)
    return json.loads(match.group(1))


def validate_summary(summary: dict[str, Any], payload: dict[str, Any] | None = None) -> list[str]:
    errors = []
    totals = summary["totals"]
    by_day = summary.get("byDay") or []
    by_material = summary.get("byMaterial") or []
    by_volunteer = summary.get("byVolunteer") or []
    latest = (payload or {}).get("latestActivity") or []

    if sum(day.get("records", 0) for day in by_day) != totals.get("records"):
        errors.append("sum(byDay.records) != totals.records")
    if sum(item.get("count", 0) for item in by_material) != totals.get("records", 0):
        errors.append("sum(byMaterial.count) != totals.records")
    for day in by_day:
        parsed = date.fromisoformat(day["dateISO"])
        if TR_WEEKDAYS[parsed.weekday()] != day.get("weekdayTR"):
            errors.append(f"weekday mismatch for {day.get('dateISO')}")
        for label in day.get("volunteerNames") or []:
            if label in {UNNAMED, HIDDEN} or is_unsafe_public_identifier(label):
                errors.append(f"non-public volunteer label in day {day.get('dateISO')}: {label}")
    busiest = max(by_day, key=lambda d: d.get("records", 0), default=None)
    highlight = (summary.get("highlights") or {}).get("busiestDay") or {}
    if busiest and highlight and busiest.get("dateISO") != highlight.get("dateISO"):
        errors.append("busiest day highlight does not match byDay")
    if summary.get("source", {}).get("recordsAreFullAggregate") is not True:
        errors.append("summary does not declare full aggregate source")
    pages_done = totals.get("pagesDone") or 0
    pages_target = totals.get("pagesTarget") or 0
    if pages_target and pages_done > pages_target:
        errors.append("pagesDone exceeds pagesTarget")
    expected_pct = round((pages_done / pages_target) * 100, 1) if pages_target else 0
    if totals.get("progressPercent") != expected_pct:
        errors.append("progress percent is not correctly rounded")
    for row in by_volunteer:
        label = row.get("label", "")
        if label in {UNNAMED, HIDDEN} or is_unsafe_public_identifier(label):
            errors.append(f"unsafe public volunteer label: {label}")
    for row in latest:
        label = row.get("volunteerLabel", "")
        if label in {UNNAMED, HIDDEN} or is_unsafe_public_identifier(label):
            errors.append(f"non-public volunteer label in latestActivity: {label}")
    for box in summary.get("byBox") or []:
        if box.get("target") and box.get("percent") != round((min(box.get("done", 0), box.get("target", 0)) / box.get("target")) * 100, 1):
            errors.append(f"active box percent mismatch: {box.get('label')}")
        if box.get("pageRows", 0) > 0 and box.get("done") is None:
            errors.append(f"active box missing done count: {box.get('label')}")
    if latest and len(latest) <= totals.get("records", 0) and summary.get("source", {}).get("latestActivityCap") != 50:
        errors.append("latestActivity cap metadata missing")
    if payload:
        text = json.dumps(payload, ensure_ascii=False)
        if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", text):
            errors.append("payload exposes an email-like value")
        if re.search(r'"volunteerToken"\s*:', text):
            errors.append("payload exposes volunteerToken")
        if re.search(r"sheet_[a-z0-9_]*row\d+", text, re.I):
            errors.append("payload exposes raw sheet row identifiers")
    return errors


def safe_sheet_title(info: dict[str, Any]) -> str:
    return info["title"]


def write_json_output(path: Path, sheet_info: list[dict[str, Any]], summaries: dict[str, Any], selected_payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    obj = {
        "generatedAt": selected_payload["generatedAt"],
        "selectedMode": selected_payload["publicSummary"]["period"]["mode"],
        "workbook": {"sheets": sheet_info},
        "summaries": summaries,
        "publicSummary": selected_payload["publicSummary"],
        "latestActivity": selected_payload["latestActivity"],
    }
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def write_report(
    path: Path,
    workbook_file: Path,
    sheet_info: list[dict[str, Any]],
    records: list[SourceRecord],
    summaries: dict[str, dict[str, Any]],
    selected_summary: dict[str, Any],
    validation_errors: list[str],
) -> None:
    calendar = summaries["calendar_week_to_date"]
    rolling = summaries["rolling_7_days"]
    page_records = [r for r in records if r.kind == "page"]
    activity_records = [r for r in records if r.kind == "activity"]
    unsafe_count = sum(1 for r in records if r.credit_status == "unsafe_identifier")
    missing_names = sum(1 for r in records if r.credit_status == "missing")

    lines = [
        "# Public Dashboard Audit",
        "",
        f"- Workbook: `{workbook_file.name}`",
        f"- Generated at: `{selected_summary['generatedAt']}`",
        f"- Selected public period: `{selected_summary['period']['label']}` (`{selected_summary['period']['mode']}`)",
        f"- Volunteer credit mode: **{selected_summary['source']['volunteerCredit']}**",
        "",
        "## Workbook Structure",
        "",
        "| Sheet | Classification | Rows | Columns | Inferred fields |",
        "|---|---:|---:|---|---|",
    ]
    for info in sheet_info:
        headers = ", ".join(info["headers"][:10])
        if len(info["headers"]) > 10:
            headers += ", ..."
        inferred = []
        for key, values in info["inferred"].items():
            if values:
                inferred.append(f"{key}: {', '.join(values[:3])}")
        lines.append(f"| {safe_sheet_title(info)} | {info['classification']} | {info['rows']} | {headers} | {'; '.join(inferred)} |")

    lines.extend([
        "",
        "## Source Counts",
        "",
        f"- Page/detail rows: **{len(page_records):,}**",
        f"- Activity rows: **{len(activity_records):,}**",
        f"- Recorded page units from detail tabs: **{sum(r.page_units for r in page_records):,}**",
        f"- Target page units from PNB inventory: **{selected_summary['totals']['pagesTarget']:,}**",
        f"- Inventory-known boxes: **{selected_summary['totals']['boxesCatalogued']:,}**",
        f"- Completed boxes: **{selected_summary['totals']['boxesCompleted']:,}**",
        "",
        "## Weekly / Rolling Window Comparison",
        "",
        "| Period | Records | Page/detail | Activity | Volunteers | Active boxes |",
        "|---|---:|---:|---:|---:|---:|",
        f"| {calendar['period']['label']} | {calendar['totals']['records']} | {calendar['totals']['pageRows']} | {calendar['totals']['activityRows']} | {calendar['totals']['volunteersActive']} | {calendar['totals']['boxesActive']} |",
        f"| {rolling['period']['label']} | {rolling['totals']['records']} | {rolling['totals']['pageRows']} | {rolling['totals']['activityRows']} | {rolling['totals']['volunteersActive']} | {rolling['totals']['boxesActive']} |",
        "",
        "## Daily Ledger",
        "",
        "| Date | Weekday | Records | Page/detail | Activity | Volunteers | Names | Boxes |",
        "|---|---|---:|---:|---:|---:|---|---|",
    ])
    for day in selected_summary["byDay"]:
        lines.append(
            f"| {day['dateISO']} | {day['weekdayTR']} | {day['records']} | {day['pageRows']} | {day['activityRows']} | {day['volunteersCount']} | {', '.join(day.get('volunteerNames') or [])} | {', '.join(day.get('boxLabels') or [])} |"
        )

    lines.extend(["", "## Material Distribution", "", "| Material | Full-period records | Share |", "|---|---:|---:|"])
    for item in selected_summary["byMaterial"]:
        lines.append(f"| {item['label']} | {item['count']} | {item['percent']}% |")

    lines.extend(["", "## Volunteer-Credit Completeness", ""])
    lines.append(f"- Public volunteer labels use real sheet names when available.")
    lines.append(f"- Rows with no usable name: **{missing_names:,}**")
    lines.append(f"- Unsafe technical identifiers suppressed: **{unsafe_count:,}**")
    lines.append("")
    lines.append("| Volunteer | Records | Page/detail | Activity | Boxes |")
    lines.append("|---|---:|---:|---:|---|")
    for row in selected_summary["byVolunteer"][:12]:
        lines.append(f"| {row['label']} | {row['records']} | {row['pageRows']} | {row['activityRows']} | {', '.join(row.get('boxes') or [])} |")

    lines.extend(["", "## Box-Progress Completeness", "", "| Box | Done | Target | Percent | Remaining | Contributors |", "|---|---:|---:|---:|---:|---|"])
    for box in selected_summary["byBox"]:
        contributors = ", ".join(f"{c['label']} +{c['records']}" for c in box.get("contributors", []))
        lines.append(f"| {box['label']} | {box['done']} | {box.get('target') or 'hedef eksik'} | {box.get('percent') if box.get('percent') is not None else '—'} | {box.get('remaining') if box.get('remaining') is not None else '—'} | {contributors} |")

    lines.extend([
        "",
        "## Detected Inconsistencies / Label Recommendations",
        "",
        "- Calendar-week and rolling-7-day windows are computed separately; the public page uses one selected period object.",
        "- Summaries are full-period aggregates; only `latestActivity` is capped and labeled as `Son 50 çalışma`.",
        "- Weekday labels are computed from `dateISO`.",
        "- Use `Envanteri girilmiş kutu`, `Tamamlanan kutu`, and `Malzeme dağılımı · bu dönem`.",
        "",
        "## Data-Quality Warnings",
        "",
    ])
    if selected_summary["warnings"]:
        for warning in selected_summary["warnings"]:
            lines.append(f"- `{warning['code']}`: {warning['message']}")
    else:
        lines.append("- None.")

    lines.extend(["", "## Validation", ""])
    if validation_errors:
        lines.extend(f"- FAIL: {err}" for err in validation_errors)
    else:
        lines.append("- PASS: summary and snapshot validation checks passed.")
    lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=None, help="Workbook path; defaults to _audit/input/*.xlsx")
    parser.add_argument("--period", choices=["calendar_week_to_date", "rolling_7_days", "calendar_week"], default="rolling_7_days")
    parser.add_argument("--write-snapshot", action="store_true", help="Write js/snapshot.js from the workbook aggregate")
    parser.add_argument("--report", type=Path, default=REPORT_PATH)
    parser.add_argument("--json", type=Path, default=SUMMARY_JSON_PATH)
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    selected_mode = "calendar_week_to_date" if args.period == "calendar_week" else args.period
    path = args.input or workbook_path()
    generated_at = datetime.now(tz=PUBLIC_TZ)
    sheet_info, rows_by_sheet = load_workbook_rows(path)
    boxes = build_inventory(rows_by_sheet)
    inventory_totals = compute_inventory_totals(rows_by_sheet)
    records = collect_records(rows_by_sheet, boxes)

    summaries = {
        "calendar_week_to_date": build_public_summary(records, boxes, inventory_totals, generated_at, "calendar_week_to_date"),
        "rolling_7_days": build_public_summary(records, boxes, inventory_totals, generated_at, "rolling_7_days"),
    }
    summary = summaries[selected_mode]
    payload = build_payload(summary, records)
    validation_errors = validate_summary(summary, payload)

    print("Workbook sheets:")
    for info in sheet_info:
        print(f"- {info['title']} ({info['classification']}, rows={info['rows']})")
    print()
    print(f"Period: {summary['period']['label']}")
    print(f"Records: {summary['totals']['records']} = {summary['totals']['pageRows']} page/detail + {summary['totals']['activityRows']} activity")
    print(f"Pages: {summary['totals']['pagesDone']} / {summary['totals']['pagesTarget']} ({summary['totals']['progressPercent']}%)")
    print(f"Boxes: {summary['totals']['boxesCatalogued']} inventory-known, {summary['totals']['boxesActive']} active this period")
    print(f"Volunteers credited: {summary['totals']['volunteersActive']}")
    print(f"Validation: {'PASS' if not validation_errors else 'FAIL'}")
    for err in validation_errors:
        print(f"  - {err}")

    if not args.validate_only:
        write_json_output(args.json, sheet_info, summaries, payload)
        write_report(args.report, path, sheet_info, records, summaries, summary, validation_errors)
        print(f"Report: {args.report}")
        print(f"JSON: {args.json}")

    if args.write_snapshot:
        write_snapshot(payload, SNAPSHOT_PATH)
        written = extract_snapshot_json(SNAPSHOT_PATH)
        written_errors = validate_summary(summary, written)
        if written_errors:
            for err in written_errors:
                print(f"Snapshot validation error: {err}")
            return 1
        print(f"Snapshot: {SNAPSHOT_PATH}")

    return 1 if validation_errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
